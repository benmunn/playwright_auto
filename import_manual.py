"""Write the manual QA notes from Level_2_tracking.xlsx into an activity workbook.

Two different shapes of note, handled two different ways:

  * Listen & Read and Listen & Read Along have no scraped content to hang off, so they get
    their own flat sheets -- one row per error, no Err[#] slots.
  * Everything else names a word, a question or an answer option that already exists in a
    scraped sheet, so it becomes an ordinary Err[#] finding with source="Manual".

Where a manual note restates something the model already found, the two are merged into
one finding marked "AI+Manual" rather than recorded twice. That is worth more than a tidy
count: it is a direct measurement of which automated checks a human agrees with.

    uv run python import_manual.py --workbook data/2plus_check.xlsx --dry-run
    uv run python import_manual.py --workbook data/2plus_check.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import manual_qa
import rs_scrape as rs

# qa_sheet owns the Err column layout; reuse its extender rather than growing a second
# implementation that could disagree about where a new set goes.
sys.path.insert(0, str(Path(__file__).parent / ".claude" / "skills" / "two-level-qa"))
import qa_sheet  # noqa: E402

LR_SHEETS = {"LR": "Listen & Read", "LRA": "Listen & Read Along"}
LR_HEADERS = ["id", "Book", "Page", "Type", "Text", "Suggested Fix", "Details", "Source"]
LR_WIDTHS = [9, 30, 7, 20, 60, 44, 40, 11]
HDR_FILL = PatternFill("solid", start_color="FFD9E1F2")
TOP = Alignment(wrap_text=True, vertical="top")

# Known typos in the hand-written notes, confirmed against the book's word list.
# `stock` is recorded for book 942 with the problem "distorted/unnatural"; the book has
# no such word but does have `stalk(s)`, i.e. the audio reads *stalk* as "stock".
WORD_ALIAS = {("942", "stock"): "stalk(s)"}

# Which cell a manual note points at, by the column it came from.
VOCAB_TARGET = {
    "Vocab - TTS": "W", "Vocab - Images": "W", "Vocab - Text": "W",
    "WMM - cut-off word": "W", "WMM - other": "W", "Vocab - Definition": "DEF",
}


def norm(text: str) -> str:
    """Loose comparison key for deciding whether two findings are the same one."""
    return re.sub(r"[^a-z0-9]+", " ", (text or "").lower()).strip()


def ensure_lr_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(LR_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment = Font(bold=True), HDR_FILL, TOP
        ws.column_dimensions[get_column_letter(c)].width = LR_WIDTHS[c - 1]
    ws.freeze_panes = "A2"
    return ws


def write_lr(wb, name: str, findings, dry_run: bool, seen: set | None = None) -> int:
    """Write one flat sheet. `seen` carries keys across sheets so LRA does not restate LR.

    The two audits were separate passes over overlapping content -- book 606's note is
    byte-identical in `LR - text` and `LRA - text` -- so without a shared key set a
    straight concatenation inflates the error count.
    """
    ws = ensure_lr_sheet(wb, name)
    existing = seen if seen is not None else set()
    existing |= {
        (str(ws.cell(r, 1).value or ""), norm(ws.cell(r, 5).value), str(ws.cell(r, 4).value or ""))
        for r in range(2, ws.max_row + 1)
    }
    written = 0
    for f in findings:
        key = (f.book_id, norm(f.text), f.type)
        if key in existing:
            continue                     # LR and LRA overlap; do not double-count
        existing.add(key)
        if not dry_run:
            row = ws.max_row + 1
            for c, v in enumerate([f.book_id, f.book, f.page, f.type, f.text, f.fix,
                                   f.details, "Manual"], 1):
                ws.cell(row, c, v).alignment = TOP
        written += 1
    return written


def head_word(word: str) -> str:
    """The word without its inflection list: `stammer, (stammered)` -> `stammer`.

    The manual notes name the bare word while the sheet carries the whole authored
    string, so an exact comparison misses most of them.
    """
    return norm(re.split(r"[,(]", word, maxsplit=1)[0])


def slot_of(ws, headers, row: int, prefix: str, word: str) -> int | None:
    """The slot whose W[#] is this word -- exact first, then ignoring inflections."""
    want, want_head = norm(word), head_word(word)
    cells = {}
    for s in range(1, rs.MAX_SLOTS + 1):
        col = headers.get(f"w{s}")
        if col:
            cells[s] = str(ws.cell(row, col).value or "")
    for s, value in cells.items():
        if norm(value) == want:
            return s
    for s, value in cells.items():
        if want_head and head_word(value) == want_head:
            return s
    return None


def cc_slot_of(ws, headers, row: int, answer: str) -> int | None:
    want = norm(answer)
    for s in range(1, rs.MAX_SLOTS + 1):
        col = headers.get(f"a{s}")
        if col and norm(str(ws.cell(row, col).value or "")) == want:
            return s
    return None


def resolve_target(ws, headers, row, f) -> str | None:
    if f.sheet == "Vocab":
        word = WORD_ALIAS.get((f.book_id, f.word.lower()), f.word)
        slot = slot_of(ws, headers, row, "W", word)
        if slot is None:
            return None
        return f"{VOCAB_TARGET.get(f.source_column, 'W')}{slot}"
    if f.sheet == "CC":
        slot = cc_slot_of(ws, headers, row, f.word)
        return f"Q{slot}" if slot else None
    return f.slot_hint or None


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    rs.WORKBOOK_PATH = args.workbook
    wb = rs.open_target_workbook()

    # Which books does this workbook actually cover? LR notes for other books belong in
    # the other batch's workbook, not copied into both.
    known: set[str] = set()
    for name in ("OEC", "CC", "Vocab", "TMC"):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        headers = rs.header_map(ws)
        known |= set(rs.sheet_book_ids(ws, headers))

    lr, lra = manual_qa.parse_listen_and_read()
    lr_mine = [f for f in lr.findings if f.book_id in known]
    lra_mine = [f for f in lra.findings if f.book_id in known]
    elsewhere = {f.book_id for f in lr.findings + lra.findings} - known
    if elsewhere:
        print(f"({len(elsewhere)} book(s) with Listen & Read notes are not in this "
              f"workbook: {', '.join(sorted(elsewhere))})")
    activities, report = manual_qa.parse_activities()
    if report.unparsed:
        print(f"!! {len(report.unparsed)} manual note(s) could not be parsed:")
        for where, entry in report.unparsed:
            print(f"   {where}: {entry[:90]!r}")

    seen: set = set()
    n_lr = write_lr(wb, "LR", lr_mine, args.dry_run, seen)
    n_lra = write_lr(wb, "LRA", lra_mine, args.dry_run, seen)
    print(f"LR : {n_lr} row(s) of {len(lr_mine)} for this workbook "
          f"({len(lr.findings)} parsed in total)")
    print(f"LRA: {n_lra} row(s) of {len(lra_mine)} for this workbook "
          f"({len(lra_mine) - n_lra} duplicated an LR entry)")

    added = merged = skipped_book = skipped_target = 0
    for f in activities:
        if f.sheet not in wb.sheetnames:
            skipped_book += 1
            continue
        ws = wb[f.sheet]
        headers = rs.header_map(ws)
        row = rs.find_row(ws, headers, f.book_id)
        if row is None:
            skipped_book += 1
            continue
        allowed = qa_sheet.SHEETS[f.sheet]["types"]
        if f.type not in allowed:
            skipped_target += 1
            print(f"  ? {f.book_id} {f.sheet}: type {f.type!r} is not allowed on this "
                  f"sheet (allowed: {', '.join(sorted(allowed))})")
            continue
        target = resolve_target(ws, headers, row, f)
        if not target or target.casefold() not in headers:
            skipped_target += 1
            # Distinguish "we never scraped this book's vocabulary" from "the word is
            # not among the ones we did scrape" -- they need different follow-up.
            empty = f.sheet == "Vocab" and not str(
                ws.cell(row, headers["w1"]).value or "").strip()
            why = ("no vocabulary was scraped for this book" if empty
                   else "not among the scraped items")
            print(f"  ? {f.book_id} {f.sheet}: cannot place "
                  f"{f.word or f.slot_hint!r} ({f.source_column}) -- {why}")
            continue

        # Does the model already have this one?
        slots = sorted({int(m.group(1)) for h in headers
                        if (m := re.fullmatch(r"err(\d+)_type", h))})
        hit = already = None
        for s in slots:
            t = str(ws.cell(row, headers[f"err{s}_target"]).value or "").strip()
            if t.casefold() != target.casefold():
                continue
            source = str(ws.cell(row, headers[f"err{s}_source"]).value or "").strip()
            same_text = norm(f.fix) and norm(f.fix) == norm(
                str(ws.cell(row, headers[f"err{s}_fix"]).value or ""))
            same_type = f.type == str(ws.cell(row, headers[f"err{s}_type"]).value or "").strip()
            if not (same_text or same_type):
                continue
            if source in ("Manual", "AI+Manual"):
                already = s     # this very note, from an earlier run -- leave it alone
                break
            hit = s
            break
        if already:
            continue
        if hit:
            merged += 1
            if not args.dry_run:
                col = headers[f"err{hit}_source"]
                ws.cell(row, col, "AI+Manual")
            continue
        added += 1
        if not args.dry_run:
            index = 1
            while True:
                key = f"err{index}_target"
                if key not in headers:
                    # Same rule the QA skill follows: extend rather than truncate.
                    qa_sheet.add_err_set(ws, headers, index)
                if not str(ws.cell(row, headers[key]).value or "").strip():
                    for field, value in (("target", target), ("type", f.type),
                                         ("details", f.details), ("fix", f.fix),
                                         ("source", "Manual")):
                        ws.cell(row, headers[f"err{index}_{field}"], value)
                    break
                index += 1

    print(f"\nactivity notes: {added} added, {merged} merged with an existing AI finding, "
          f"{skipped_book} for books not in this workbook, {skipped_target} unplaceable")
    if args.dry_run:
        print("\nDRY RUN -- nothing written.")
    else:
        rs.save_workbook(wb)
        print(f"\nSaved {args.workbook}")


if __name__ == "__main__":
    main()

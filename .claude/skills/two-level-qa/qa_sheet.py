"""Spreadsheet mechanics for the two-level QA skill.

The model does the judging; this script does everything deterministic around it, so a
long QA run cannot lose work or silently corrupt the source data:

    status    what remains to judge, per sheet
    prepare   back up the workbook, then append the Err column headers
    batch     emit the next N unjudged rows as JSON
    write     record findings, extending Err6+ when a row needs them
    proof     emit judged rows together with their recorded findings, for pass 2
    retract   remove a finding pass 2 rejected, closing the gap behind it

Header lookup and id normalisation are imported from rs_scrape rather than reimplemented,
so a row this script calls 697 is the same row the scraper wrote.

Two bookkeeping columns beyond the 20 the instructions specify are unavoidable: a row
with no errors is otherwise indistinguishable from a row nobody has looked at yet.
QA_pass1 and QA_pass2 hold the finding count for that pass ("0" for a clean row), which
is what makes an interrupted run resumable.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from rs_scrape import (  # noqa: E402
    ID_HEADER,
    SheetLayoutError,
    _cell_key,
    column_for,
    header_map,
)

ERR_FIELDS = ("target", "type", "details", "fix", "source")
BASE_ERR_SETS = 5
PASS_COLUMNS = ("QA_pass1", "QA_pass2")

# Who found the error. Findings imported from a human pass are worth telling apart from
# the model's: it says which detectors a person has already confirmed, and "AI+Manual" on
# a row is a direct measurement that an automated check agrees with a human one.
SOURCES = ("AI", "Manual", "AI+Manual")
DEFAULT_SOURCE = "AI"

ERR_FILL = PatternFill("solid", start_color="FFFCE4D6")
ERR_FONT = Font(bold=True)


# --------------------------------------------------------------------------------------
# Per-sheet definitions. `items` describes how one row breaks into judgeable units;
# `types` is the closed set of Err_type values that sheet accepts.
#
# Every sheet accepts "Too Hard": the source instructions offered it only for OEC,
# but above-level vocabulary turns up in all four activities and needs somewhere to go.
#
# The taxonomies also correct two copy-paste faults in the source instructions: OEC's allowed
# list is the one its own body defines (not CC's), and Vocab's part-of-speech type is
# called "Part of Speech" (the doc's Types line said "Inflection"). See
# reference/error-types.md.
# --------------------------------------------------------------------------------------

COMMON = ("Grammar", "Punctuation", "Capitalization", "Spacing", "Awkward Phrasing")

SHEETS: dict[str, dict] = {
    "OEC": {
        "unit": ("Q",),
        "anchor": "Q",
        # Flat, one-per-book columns carried into every batch as context. The student
        # sees both while answering, so they are what "Requires Reading" is judged
        # against -- and they are judgeable content in their own right. Populated by
        # book_fields.py; null in any workbook scraped before it existed.
        "context": ("Main_Character", "Prev_Text"),
        "types": (*COMMON, "Unclear", "Requires Reading", "Too Hard", "Other"),
    },
    "CC": {
        "unit": ("Q", "A"),
        "anchor": "Q",
        "types": (*COMMON, "Lack of Context", "Answer Given", "Too Hard", "Other"),
    },
    "Vocab": {
        # The whole Vocab sheet comes from the book's word list at /books/<id>/words.
        # SENT is the sentence showing how this book uses the word -- the evidence for
        # the two types below, which were previously inferred from the book's other
        # vocabulary, so it travels in the batch payload though students never see it.
        "unit": ("W", "POS", "DEF", "SENT"),
        "anchor": "W",
        # "Wrong Sense" covers a definition that is accurate for some other meaning of
        # the word but not the one this book teaches -- "hemisphere" glossed as half a
        # brain in a book about continents. It was previously logged as "Too Hard",
        # which sent the fixer looking for a simpler word instead of a different sense.
        # "Story Sentence" is for a fault in SENT itself -- a run-together after a full
        # stop, a garbled clause -- as opposed to a fault the sentence reveals in the
        # definition. Students never see SENT, so these are lower priority than the
        # rest and worth being able to filter out.
        "types": (*COMMON, "Part of Speech", "Wrong Sense", "Lack of Context",
                  "Answer Given", "Too Hard", "Story Sentence", "Other"),
    },
    "TMC": {
        "unit": ("Q", "AnsA", "AnsB", "AnsC", "AnsD", "Correct"),
        "anchor": "Q",
        "types": (*COMMON, "Unclear", "Too Hard", "Other"),
    },
}

MAX_SLOTS = 30


def spec(sheet: str) -> dict:
    if sheet not in SHEETS:
        raise SystemExit(f"Unknown sheet {sheet!r}. Expected one of: {', '.join(SHEETS)}")
    return SHEETS[sheet]


def valid_target(sheet: str, headers: dict[str, int], target: str) -> bool:
    """Is `target` something a finding can point at on this sheet?

    Normally a target must be a real column. TMC adds one virtual target: `AnsAll[#]`
    means all four options of question #, for a fault they share -- periods on every
    option, or a whole option set pasted from the wrong book. Without it the same
    finding has to be written four times, and a fixer reading four near-identical rows
    cannot tell whether they are one problem or four.
    """
    if target.casefold() in headers:
        return True
    if sheet == "TMC":
        m = re.fullmatch(r"AnsAll(\d+)", target, re.I)
        if m and f"ansa{m.group(1)}" in headers:
            return True
    return False


# --------------------------------------------------------------------------------------
# Workbook access
# --------------------------------------------------------------------------------------


def open_workbook(path: Path):
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")
    lock = path.with_name("~$" + path.name)
    if lock.exists():
        raise SystemExit(
            f"{path.name} is open in Excel ({lock.name} present).\n"
            "Close it first -- openpyxl cannot write to a locked file, and a QA run "
            "that fails on save would lose the whole batch."
        )
    return load_workbook(path)


def save(wb, path: Path) -> None:
    try:
        wb.save(path)
    except PermissionError:
        raise SystemExit(
            f"Cannot save {path} -- it is probably open in Excel. Close it and re-run; "
            "rows already recorded will be skipped."
        )


def err_set_count(headers: dict[str, int]) -> int:
    """How many complete Err[#] column sets the sheet currently has."""
    count = 0
    while all(
        f"err{count + 1}_{field}".casefold() in headers for field in ERR_FIELDS
    ):
        count += 1
    return count


def add_err_set(ws: Worksheet, headers: dict[str, int], index: int) -> None:
    """Append one Err[#] set at the end of the sheet and register it in `headers`."""
    for field in ERR_FIELDS:
        name = f"Err{index}_{field}"
        column = ws.max_column + 1
        cell = ws.cell(row=1, column=column, value=name)
        cell.fill = ERR_FILL
        cell.font = ERR_FONT
        ws.column_dimensions[get_column_letter(column)].width = (
            14 if field == "target" else 20 if field == "type" else 46
        )
        headers[name.casefold()] = column


def content_slots(ws: Worksheet, headers: dict[str, int], row: int, sheet: str) -> list[int]:
    """Slot numbers 1..MAX_SLOTS that actually hold content in this row."""
    anchor = spec(sheet)["anchor"]
    slots = []
    for slot in range(1, MAX_SLOTS + 1):
        key = f"{anchor}{slot}".casefold()
        if key not in headers:
            break
        if _cell_key(ws.cell(row=row, column=headers[key]).value):
            slots.append(slot)
    return slots


def data_rows(ws: Worksheet, headers: dict[str, int], sheet: str) -> list[int]:
    """Rows carrying an id and at least one populated content cell."""
    id_column = column_for(ws, headers, ID_HEADER)
    rows = []
    for row in range(2, ws.max_row + 1):
        if not _cell_key(ws.cell(row=row, column=id_column).value):
            continue
        if content_slots(ws, headers, row, sheet):
            rows.append(row)
    return rows


def pass_state(ws: Worksheet, headers: dict[str, int], row: int, which: int) -> str:
    key = PASS_COLUMNS[which - 1].casefold()
    if key not in headers:
        return ""
    return _cell_key(ws.cell(row=row, column=headers[key]).value)


def read_findings(ws: Worksheet, headers: dict[str, int], row: int) -> list[dict]:
    """Every recorded finding in this row, each tagged with its slot number."""
    out = []
    for index in range(1, err_set_count(headers) + 1):
        entry = {
            field: _cell_key(
                ws.cell(row=row, column=headers[f"err{index}_{field}".casefold()]).value
            )
            for field in ERR_FIELDS
        }
        if any(entry.values()):
            out.append({"slot": index, **entry})
    return out


# --------------------------------------------------------------------------------------
# Commands
# --------------------------------------------------------------------------------------


def cmd_status(args) -> None:
    wb = open_workbook(args.workbook)
    print(f"{args.workbook}\n")
    print(f"{'sheet':6} {'rows':>6} {'pass1':>7} {'pass2':>7} {'findings':>9} {'Err sets':>9}")
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f"{sheet:6} {'-- not in workbook --':>41}")
            continue
        ws = wb[sheet]
        headers = header_map(ws)
        rows = data_rows(ws, headers, sheet)
        p1 = sum(1 for r in rows if pass_state(ws, headers, r, 1))
        p2 = sum(1 for r in rows if pass_state(ws, headers, r, 2))
        found = sum(len(read_findings(ws, headers, r)) for r in rows)
        print(f"{sheet:6} {len(rows):6} {p1:7} {p2:7} {found:9} {err_set_count(headers):9}")


def cmd_prepare(args) -> None:
    wb = open_workbook(args.workbook)
    backup = args.workbook.with_name(
        f"{args.workbook.stem}.backup-{time.strftime('%Y%m%d-%H%M%S')}{args.workbook.suffix}"
    )
    if not args.no_backup:
        shutil.copy2(args.workbook, backup)
        print(f"backup: {backup.name}")

    for sheet in args.sheets:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"workbook has no {sheet!r} sheet")
        ws = wb[sheet]
        headers = header_map(ws)
        column_for(ws, headers, ID_HEADER)  # fail loudly if the layout is wrong

        added = 0
        for index in range(err_set_count(headers) + 1, BASE_ERR_SETS + 1):
            add_err_set(ws, headers, index)
            added += len(ERR_FIELDS)
        for name in PASS_COLUMNS:
            if name.casefold() not in headers:
                column = ws.max_column + 1
                cell = ws.cell(row=1, column=column, value=name)
                cell.fill = ERR_FILL
                cell.font = ERR_FONT
                ws.column_dimensions[get_column_letter(column)].width = 10
                headers[name.casefold()] = column
                added += 1
        print(f"{sheet}: {added} column(s) added, {err_set_count(headers)} Err set(s)")

    save(wb, args.workbook)


def cmd_batch(args) -> None:
    wb = open_workbook(args.workbook)
    sheet = args.sheet
    ws = wb[sheet]
    headers = header_map(ws)
    which = args.qa_pass
    id_column = column_for(ws, headers, ID_HEADER)

    payload = {"sheet": sheet, "pass": which, "allowed_types": list(spec(sheet)["types"]),
               "rows": []}
    for row in data_rows(ws, headers, sheet):
        if pass_state(ws, headers, row, which):
            continue
        if which == 2 and not pass_state(ws, headers, row, 1):
            continue  # never proof a row pass 1 has not seen
        entry = {
            "row": row,
            "id": _cell_key(ws.cell(row=row, column=id_column).value),
            "items": [],
        }
        for slot in content_slots(ws, headers, row, sheet):
            item = {"slot": slot}
            for prefix in spec(sheet)["unit"]:
                key = f"{prefix}{slot}".casefold()
                # null means the column does not exist at all -- distinct from "", an
                # empty cell. TMC's Correct[#] is null until the sheet is re-scraped
                # with the answer key, and QA must not silently assume it was checked.
                item[f"{prefix}{slot}"] = (
                    _cell_key(ws.cell(row=row, column=headers[key]).value)
                    if key in headers
                    else None
                )
            entry["items"].append(item)

        # Flat per-book context columns (OEC's Main_Character / Prev_Text). Same null
        # rule as the item cells above: null means the column is absent from the sheet,
        # "" means the book left it blank. The distinction matters -- a workbook that
        # predates book_fields.py must not read as "checked and empty".
        for name in spec(sheet).get("context", ()):
            key = name.casefold()
            entry[name] = (
                _cell_key(ws.cell(row=row, column=headers[key]).value)
                if key in headers
                else None
            )

        if sheet == "CC":
            # Judging "lack of context" and "answer given" needs the whole word bank,
            # not just this sentence's own answer.
            entry["word_bank"] = [
                _cell_key(ws.cell(row=row, column=headers[f"a{i}".casefold()]).value)
                for i in range(1, MAX_SLOTS + 1)
                if f"a{i}".casefold() in headers
                and _cell_key(ws.cell(row=row, column=headers[f"a{i}".casefold()]).value)
            ]
        if which == 2:
            entry["recorded"] = read_findings(ws, headers, row)

        payload["rows"].append(entry)
        if len(payload["rows"]) >= args.size:
            break

    # The sheets are full of curly quotes, and this console is cp1252: dumping straight
    # to stdout would emit bytes that are not valid UTF-8. Always name the encoding.
    text = json.dumps(payload, ensure_ascii=False, indent=1)
    if args.out:
        args.out.write_text(text, encoding="utf-8")
        print(
            f"{sheet} pass {which}: {len(payload['rows'])} row(s) -> {args.out} "
            f"({sum(len(r['items']) for r in payload['rows'])} items to judge)"
        )
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        print(text)


def cmd_write(args) -> None:
    wb = open_workbook(args.workbook)
    sheet = args.sheet
    ws = wb[sheet]
    headers = header_map(ws)
    allowed = set(spec(sheet)["types"])
    which = args.qa_pass

    data = json.loads(args.findings.read_text(encoding="utf-8"))
    rows = data["rows"] if isinstance(data, dict) else data

    # Validate everything before writing anything, so a bad batch cannot half-apply.
    problems = []
    for entry in rows:
        for finding in entry.get("findings", []):
            if finding.get("type") not in allowed:
                problems.append(
                    f"row {entry['row']}: type {finding.get('type')!r} not allowed for "
                    f"{sheet} (allowed: {', '.join(sorted(allowed))})"
                )
            target = (finding.get("target") or "").strip()
            if not valid_target(sheet, headers, target):
                problems.append(
                    f"row {entry['row']}: target {target!r} is not a column in {sheet}"
                )
            if not (finding.get("details") or "").strip():
                problems.append(f"row {entry['row']}: empty details")
            source = (finding.get("source") or DEFAULT_SOURCE).strip()
            if source not in SOURCES:
                problems.append(
                    f"row {entry['row']}: source {source!r} is not one of "
                    f"{', '.join(SOURCES)}"
                )
    if problems:
        raise SystemExit("Refusing to write:\n  " + "\n  ".join(problems))

    written = extended = 0
    for entry in rows:
        row = entry["row"]
        findings = entry.get("findings", [])
        for finding in findings:
            index = 1
            while True:
                if index > err_set_count(headers):
                    add_err_set(ws, headers, index)
                    extended += 1
                target_col = headers[f"err{index}_target".casefold()]
                if not _cell_key(ws.cell(row=row, column=target_col).value):
                    break
                index += 1
            for field in ERR_FIELDS:
                default = DEFAULT_SOURCE if field == "source" else ""
                ws.cell(
                    row=row,
                    column=headers[f"err{index}_{field}".casefold()],
                    value=(finding.get(field) or default).strip(),
                )
            written += 1

        stamp_col = headers.get(PASS_COLUMNS[which - 1].casefold())
        if stamp_col is None:
            raise SheetLayoutError(
                f"sheet {sheet!r} has no {PASS_COLUMNS[which - 1]} column -- run prepare"
            )
        ws.cell(row=row, column=stamp_col, value=len(findings))

    save(wb, args.workbook)
    print(
        f"{sheet} pass {which}: {len(rows)} row(s) marked judged, {written} finding(s) "
        f"recorded, {extended} new Err column set(s)"
    )


def cmd_retract(args) -> None:
    wb = open_workbook(args.workbook)
    ws = wb[args.sheet]
    headers = header_map(ws)
    data = json.loads(args.retractions.read_text(encoding="utf-8"))
    entries = data["retractions"] if isinstance(data, dict) else data

    by_row: dict[int, set[int]] = {}
    for item in entries:
        by_row.setdefault(item["row"], set()).add(item["slot"])

    removed = 0
    for row, slots in by_row.items():
        kept = [f for f in read_findings(ws, headers, row) if f["slot"] not in slots]
        removed += len(read_findings(ws, headers, row)) - len(kept)
        # Rewrite the row's findings from slot 1 so no gap is left behind. Assign to
        # .value rather than passing value= to cell(): openpyxl ignores a None passed
        # that way, which would leave the vacated tail slots holding stale findings.
        for index in range(1, err_set_count(headers) + 1):
            source = kept[index - 1] if index <= len(kept) else None
            for field in ERR_FIELDS:
                cell = ws.cell(row=row, column=headers[f"err{index}_{field}".casefold()])
                cell.value = source[field] if source else None
        # Update the count for the pass doing the retracting. Defaulting to pass 2
        # unconditionally would mark a row proofed when a pass-1 correction is what
        # actually happened, and the row would then be skipped by the proof pass.
        stamp = headers.get(PASS_COLUMNS[args.qa_pass - 1].casefold())
        if stamp:
            ws.cell(row=row, column=stamp, value=len(kept))

    save(wb, args.workbook)
    print(f"{args.sheet}: {removed} finding(s) retracted across {len(by_row)} row(s)")


# --------------------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = parser.add_subparsers(dest="command", required=True)

    def add(name, fn, **kw):
        p = sub.add_parser(name, **kw)
        p.add_argument("workbook", type=Path)
        p.set_defaults(func=fn)
        return p

    add("status", cmd_status, help="what remains to judge")

    p = add("prepare", cmd_prepare, help="backup, then add Err columns")
    p.add_argument("--sheets", nargs="+", default=list(SHEETS))
    p.add_argument("--no-backup", action="store_true")

    p = add("batch", cmd_batch, help="emit the next unjudged rows as JSON")
    p.add_argument("--sheet", required=True)
    p.add_argument("--size", type=int, default=15)
    p.add_argument("--pass", dest="qa_pass", type=int, choices=(1, 2), default=1)
    p.add_argument(
        "--out", type=Path,
        help="write the JSON here (recommended -- avoids console encoding problems)",
    )

    p = add("write", cmd_write, help="record findings")
    p.add_argument("--sheet", required=True)
    p.add_argument("--findings", type=Path, required=True)
    p.add_argument("--pass", dest="qa_pass", type=int, choices=(1, 2), default=1)

    p = add("retract", cmd_retract, help="remove findings rejected by pass 2")
    p.add_argument("--sheet", required=True)
    p.add_argument("--retractions", type=Path, required=True)
    p.add_argument(
        "--pass", dest="qa_pass", type=int, choices=(1, 2), default=2,
        help="which pass is retracting; use 1 when correcting during pass 1",
    )

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()

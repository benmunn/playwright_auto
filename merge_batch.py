"""Append one activity workbook's rows into another and mark what was appended.

The two workbooks were scraped as separate batches and hold disjoint book ids, so a
merge is a straight append rather than a reconciliation -- there is nothing to resolve.
Every appended row is filled with a colour so the reviewer can see at a glance which
books arrived from the second batch and still want their eyes.

    uv run python merge_batch.py --into data/2plus_check.xlsx \\
        --from data/2plus_check_batch2.xlsx --dry-run
    uv run python merge_batch.py --into data/2plus_check.xlsx \\
        --from data/2plus_check_batch2.xlsx

Refuses to run if a book id appears in both files: that would mean the two are versions
of the same rows, and appending would silently double them.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Lavender, which is not one of the report legend's colours and not one of the two the
# change highlighter already uses, so "merged from the second batch" reads as its own
# thing in the workbook and in the reports alike.
MERGED_FILL = PatternFill("solid", start_color="FFE4D7F5")
NOTE_FILL = PatternFill("solid", start_color="FFE4D7F5")


def headers(ws) -> dict[str, int]:
    """Header name -> column index, so columns are matched by name, not position."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            out[str(v).strip()] = c
    return out


def ids_in(ws, hdr: dict[str, int]) -> list:
    col = hdr.get("id")
    if col is None:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip():
            out.append(str(int(v)) if isinstance(v, float) and v.is_integer()
                       else str(v).strip())
    return out


def data_rows(ws, hdr: dict[str, int]) -> list[int]:
    """Row numbers that hold data, judged by any non-empty cell rather than by `id`.

    The flat sheets key on `id` too, but a row with an id and nothing else is still a
    row worth carrying over, and a blank trailing row is not.
    """
    col = hdr.get("id")
    out = []
    for r in range(2, ws.max_row + 1):
        if col is not None:
            v = ws.cell(r, col).value
            if v is None or not str(v).strip():
                continue
            out.append(r)
        elif any(ws.cell(r, c).value is not None
                 for c in range(1, ws.max_column + 1)):
            out.append(r)
    return out


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--into", type=Path, required=True)
    ap.add_argument("--from", dest="src", type=Path, required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-highlight", action="store_true")
    args = ap.parse_args()

    dst_wb = load_workbook(args.into)
    src_wb = load_workbook(args.src, data_only=True)

    missing = [s for s in src_wb.sheetnames if s not in dst_wb.sheetnames]
    if missing:
        sys.exit(f"{args.into.name} has no {', '.join(missing)} sheet(s) to merge into.")

    # An overlapping id would mean these are two versions of the same rows.
    clashes: dict[str, list] = {}
    for name in src_wb.sheetnames:
        d, s = dst_wb[name], src_wb[name]
        both = set(ids_in(d, headers(d))) & set(ids_in(s, headers(s)))
        if both:
            clashes[name] = sorted(both)
    if clashes:
        for name, ids in clashes.items():
            print(f"!! {name}: {len(ids)} id(s) in both files: {', '.join(ids[:10])}")
        sys.exit("Refusing to merge -- appending would duplicate these rows.")

    total_rows = total_cells = 0
    for name in src_wb.sheetnames:
        dst, src = dst_wb[name], src_wb[name]
        d_hdr, s_hdr = headers(dst), headers(src)
        unknown = [h for h in s_hdr if h not in d_hdr]
        if unknown:
            print(f"!! {name}: {len(unknown)} column(s) exist only in "
                  f"{args.src.name} and would be dropped: {', '.join(unknown[:8])}")
            sys.exit("Refusing to merge -- extend the target's columns first.")

        rows = data_rows(src, s_hdr)
        first_new = dst.max_row + 1
        for i, r in enumerate(rows):
            out_row = first_new + i
            for h, sc in s_hdr.items():
                value = src.cell(r, sc).value
                if value is None:
                    continue
                cell = dst.cell(out_row, d_hdr[h])
                cell.value = value
                total_cells += 1
            if not args.no_highlight:
                for c in range(1, dst.max_column + 1):
                    dst.cell(out_row, c).fill = MERGED_FILL
        total_rows += len(rows)
        print(f"  {name:12} +{len(rows):>4} row(s) "
              f"(rows {first_new}-{first_new + len(rows) - 1})" if rows
              else f"  {name:12} +   0 row(s)")

    if not args.no_highlight:
        # A legend on the first sheet, off to the right of the Err block, so the colour
        # means something to whoever opens the file next week.
        ws = dst_wb[dst_wb.sheetnames[0]]
        c = ws.max_column + 2
        ws.cell(1, c, f"Merged from {args.src.name}").font = Font(bold=True)
        ws.cell(1, c).fill = NOTE_FILL
        ws.cell(2, c, "Highlighted rows came from the second scrape batch and have "
                      "not been reviewed by hand in this file yet.")
        ws.column_dimensions[get_column_letter(c)].width = 46

    print(f"\n{total_rows} row(s), {total_cells} cell(s) appended")
    if args.dry_run:
        print("DRY RUN -- nothing saved.")
        return
    dst_wb.save(args.into)
    print(f"saved {args.into}")


if __name__ == "__main__":
    main()

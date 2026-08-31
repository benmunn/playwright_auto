"""Build the three error-report workbooks from a QA'd activity workbook.

    uv run python make_reports.py --workbook data/2plus_check.xlsx \
        --out-prefix references/0821_error-reports

Writes three views of the same findings -- one row per error, grouped by recurring kind,
and grouped by the type each was logged under -- plus a consistency check that all three
hold the same rows. Book titles come from the tracking file so the reports name books
rather than ids.
"""

from __future__ import annotations

import argparse
import collections
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

import manual_qa
from report.classify import classify
from report.plain import BANNED, plain
from report.render import render_by_type, render_long, render_recurring

SHEETS = ("OEC", "CC", "Vocab", "TMC")
FLAT_SHEETS = ("LR", "LRA")
ACTIVITY = {
    "OEC": ("Open-Ended Questions", "Let's Explore"),
    "CC": ("Context Clue", "After Reading"),
    "Vocab": ("Word Meaning Match", "Before Reading"),
    "TMC": ("Text Multiple Choice", "After Reading"),
    "LR": ("Listen & Read", "Reading"),
    "LRA": ("Listen & Read Along", "Reading"),
}

# The audit recorded a type and a quoted sentence but usually no prose, so each type
# supplies the sentence a reader of the report needs and the action it implies.
LR_DETAILS = {
    "Severe TTS Cutoff": ("the recorded audio stops well before the end of this "
                          "sentence", "Re-record the audio for this sentence."),
    "Minor TTS Cutoff": ("the recorded audio is clipped slightly at the end of this "
                         "sentence", "Re-record the audio for this sentence."),
    "Start TTS Cutoff": ("the recorded audio starts late, so the beginning of this "
                         "sentence is missing",
                         "Re-record the audio for this sentence."),
    "TTS Pronunciation": ("the recorded audio mispronounces a word in this sentence",
                          "Re-record the audio for this sentence."),
    "Text-TTS Mismatch": ("the recorded audio does not say what the page shows",
                          "Make the audio and the printed text say the same thing."),
    "Text": ("there is a mistake in the printed text on this page",
             "Correct the text on the page."),
    "Other": ("a reviewer marked a problem on this page",
              "Check this page and correct the problem described."),
}


def load_flat_findings(wb, titles: dict[str, str]) -> list[dict]:
    """Listen & Read findings, which are flat rows rather than Err[#] sets.

    These activities have no scraped content to hang a finding off -- there is no
    question or word list, only a page and the sentence the reviewer quoted -- so they
    live on their own sheets and are lifted into the same finding shape here.
    """
    out: list[dict] = []
    for sheet in FLAT_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = {str(ws.cell(1, c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        act, phase = ACTIVITY[sheet]
        for rn in range(2, ws.max_row + 1):
            def g(name: str) -> str:
                c = hdr.get(name)
                v = ws.cell(rn, c).value if c else None
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                return "" if v is None else str(v).strip()

            bid, kind_type = g("id"), g("Type")
            if not bid or not kind_type:
                continue
            page = g("Page")
            default_details, default_fix = LR_DETAILS.get(kind_type, ("", ""))
            raw = g("Details") or default_details
            out.append({
                "sheet": sheet, "book_id": bid,
                "book": g("Book") or titles.get(bid, ""),
                "activity": act, "phase": phase,
                "target": f"Page {page}" if page else "Page not recorded",
                "prefix": "Page", "n": int(page) if page.isdigit() else None,
                "type": kind_type, "details_raw": raw, "details": plain(raw),
                "fix": g("Suggested Fix") or default_fix, "current": g("Text"),
                "ctx": {"page": page, "text": g("Text")},
                "source": g("Source") or "Manual",
            })
    return out


def load_glossary() -> dict[str, dict]:
    """The global word record for each id, which is what carries the translations.

    Absent cache is not fatal: the report still builds, the translation columns are just
    empty, and the run says so rather than failing.
    """
    import word_ids

    if not word_ids.CACHE.exists():
        print(f"!! {word_ids.CACHE} is missing -- the translation columns will be blank. "
              f"Run: uv run python word_ids.py --scrape")
        return {}
    glossary = {w["id"]: w for w in word_ids.load_global()}
    if not any("kor" in w for w in glossary.values()):
        print(f"!! {word_ids.CACHE} predates the translation columns -- re-scrape to "
              f"fill them.")
    return glossary


def load_word_uses(path: Path) -> dict[str, list[dict]]:
    """word id -> every book that links to it, flagged or not.

    The findings alone only show the entries somebody complained about; deciding whether
    one edit can serve every book needs the quiet uses too.
    """
    ws = load_workbook(path, data_only=True)["Vocab"]
    hdr = {str(ws.cell(1, c).value).strip(): c
           for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    out: dict[str, list[dict]] = collections.defaultdict(list)
    for rn in range(2, ws.max_row + 1):
        bid = ws.cell(rn, hdr["id"]).value
        if bid is None:
            continue
        bid = str(int(bid)) if isinstance(bid, float) else str(bid).strip()
        for s in range(1, 26):
            if f"W{s}" not in hdr:
                break
            cell = lambda name: str(ws.cell(rn, hdr[name]).value or "").strip()
            wid = ws.cell(rn, hdr[f"WID{s}"]).value if f"WID{s}" in hdr else None
            if not cell(f"W{s}") or not wid:
                continue
            out[str(int(wid))].append({
                "book": bid, "word": cell(f"W{s}"), "pos": cell(f"POS{s}"),
                "definition": cell(f"DEF{s}"), "sentence": cell(f"SENT{s}")})
    _apply_journal(out)
    return out


def _apply_journal(uses: dict[str, list[dict]]) -> None:
    """Bring the current-value columns up to date with edits already applied.

    The workbook records what each entry said when it was scraped. Once word_edit.py
    has written to the live list those columns describe the past, and anything that
    compares them against the site -- the identity check that guards every edit --
    reads the difference as the entry having drifted and refuses to touch it. The
    journal is the record of what was written, so it is replayed over them here.
    """
    journal = Path("data/word_edits.jsonl")
    if not journal.exists():
        return
    field = {"word": "word", "pos": "pos", "definition": "definition"}
    for line in journal.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        if rec.get("status") != "saved":
            continue
        for use in uses.get(str(rec["word_id"]), []):
            for f in rec.get("fields", []):
                if f in field:
                    use[field[f]] = rec["after"][f]


def load_findings(path: Path, titles: dict[str, str]) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    out: list[dict] = []
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = {str(ws.cell(1, c).value).strip(): c
               for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        slots = sorted({int(m.group(1)) for h in hdr
                        if (m := re.fullmatch(r"Err(\d+)_type", h))})

        def g(rn: int, name: str) -> str:
            c = hdr.get(name)
            v = ws.cell(rn, c).value if c else None
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            return "" if v is None else str(v).strip()

        for rn in range(2, ws.max_row + 1):
            bid = g(rn, "id")
            if not bid:
                continue
            for s in slots:
                kind_type = g(rn, f"Err{s}_type")
                if not kind_type:
                    continue
                tgt = g(rn, f"Err{s}_target")
                m = re.fullmatch(r"([A-Za-z_]+?)(\d+)", tgt)
                prefix, n = (m.group(1), int(m.group(2))) if m else (tgt, None)
                ctx: dict = {}
                if sheet == "TMC" and n:
                    ctx = {"question": g(rn, f"Q{n}"), "key": g(rn, f"Correct{n}"),
                           "options": {p: g(rn, f"Ans{p}{n}") for p in "ABCD"}}
                elif sheet == "CC" and n:
                    ctx = {"sentence": g(rn, f"Q{n}"), "answer": g(rn, f"A{n}")}
                elif sheet == "Vocab" and n:
                    # WID is the global word list's id for this entry, which is what
                    # tells two books' uses of the same spelling apart.
                    ctx = {"word": g(rn, f"W{n}"), "pos": g(rn, f"POS{n}"),
                           "definition": g(rn, f"DEF{n}"), "sentence": g(rn, f"SENT{n}"),
                           "word_id": g(rn, f"WID{n}")}
                elif sheet == "OEC":
                    ctx = {"main_character": g(rn, "Main_Character"),
                           "prev_text": g(rn, "Prev_Text")}
                # An AnsAll row shows every option, both as authored and as fixed.
                current = ("\n".join(f"{p}. {g(rn, f'Ans{p}{n}')}" for p in "ABCD")
                           if prefix == "AnsAll" and n else g(rn, tgt))
                raw = g(rn, f"Err{s}_details")
                act, phase = ACTIVITY[sheet]
                # The story sentence lives on the book's word list, not in Word Meaning
                # Match -- students never see it there -- so it is named for the page
                # whoever fixes it has to open.
                if sheet == "Vocab" and kind_type == "Story Sentence":
                    act = "Book Word List"
                out.append({
                    "sheet": sheet, "book_id": bid, "book": titles.get(bid, ""),
                    "activity": act, "phase": phase, "target": tgt, "prefix": prefix,
                    "n": n, "type": kind_type, "details_raw": raw, "details": plain(raw),
                    "fix": g(rn, f"Err{s}_fix"), "current": current, "ctx": ctx,
                    "source": g(rn, f"Err{s}_source") or "AI",
                })
    out += load_flat_findings(wb, titles)
    for f in out:
        # "Wrong Sense" is its own logged type but shares a recurring kind with the
        # findings still logged as "Too Hard" that describe the same fault.
        f["kind"] = ("W02" if f["sheet"] == "Vocab" and f["type"] == "Wrong Sense"
                     else classify({"sheet": f["sheet"], "type": f["type"],
                                    "details": f["details_raw"]}))
    return out


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--out-prefix", required=True,
                    help="e.g. references/0821_error-reports")
    args = ap.parse_args()

    _, rows = manual_qa.load_books()
    titles = {b: t for b, t, _ in rows}

    data = load_findings(args.workbook, titles)
    if not data:
        sys.exit(f"no findings in {args.workbook}")

    unclassified = [f for f in data if f["kind"] == "UNCLASSIFIED"]
    if unclassified:
        print(f"!! {len(unclassified)} finding(s) matched no recurring-error rule:")
        for f in unclassified[:10]:
            print(f"   {f['sheet']} {f['book_id']} {f['type']}: {f['details'][:60]}")
        sys.exit("\nRefusing to render: every finding needs a recurring-error rule, or "
                 "the by-kind and by-type views would disagree with the long one. Add a "
                 "rule in report/classify.py, or correct the finding's type.")
    jargon = [f for f in data if BANNED.search(f["details"])]
    if jargon:
        print(f"!! {len(jargon)} finding(s) still contain jargon after rewriting")
    missing = {f["book_id"] for f in data if not f["book"]}
    if missing:
        print(f"!! {len(missing)} book id(s) have no title in the tracking file: "
              f"{', '.join(sorted(missing)[:8])}")

    paths = [
        render_long(data, f"{args.out_prefix}_long.xlsx"),
        render_recurring(data, f"{args.out_prefix}_recurring.xlsx"),
        render_by_type(data, f"{args.out_prefix}_by-type.xlsx",
                       load_word_uses(args.workbook), titles, load_glossary()),
    ]
    print(f"\n{len(data)} finding(s) across {len({f['book_id'] for f in data})} book(s)")
    print("by source :", dict(collections.Counter(f["source"] for f in data)))
    print("by sheet  :", dict(collections.Counter(f["sheet"] for f in data)))
    for p in paths:
        print(f"  wrote {p}")

    # The three views must agree; that is the whole point of generating them together.
    key = lambda f: (f["book_id"], f["target"], f["type"], f["details"], f["fix"])
    want = collections.Counter(key(f) for f in data)
    for p in paths[1:]:
        wb = load_workbook(p, data_only=True)
        got: collections.Counter = collections.Counter()
        for name in wb.sheetnames[1:]:
            ws = wb[name]
            idx = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)
                   if ws.cell(4, c).value}
            # A finding detail sheet is one with a Field column. The consolidated vocab
            # sheets carry a Book ID too, but their rows are word entries, not findings,
            # so counting them here would break the very check this performs.
            if "Field" not in idx:
                continue
            for r in range(5, ws.max_row + 1):
                cell = lambda h: ("" if idx.get(h) is None
                                  or ws.cell(r, idx[h]).value is None
                                  else str(ws.cell(r, idx[h]).value).strip())
                if cell("Book ID"):
                    got[(cell("Book ID"), cell("Field"), cell("Reason"),
                         cell("Details"), cell("Suggested Fix"))] += 1
        print(f"  {p.name} matches the source rows: {got == want}")


if __name__ == "__main__":
    main()

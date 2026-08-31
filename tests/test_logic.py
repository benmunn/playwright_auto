"""Tests for the decision logic behind the reports and the edit run.

Nothing here opens a browser or touches the network. What is covered is the reasoning a
run depends on -- which rows are held back, which edits the database would refuse,
whether a word still has work outstanding -- because that is where the faults have
actually been. Each test below stands for a bug that reached the live word list or came
within one run of doing so.

    uv run python tests/test_logic.py
"""

from __future__ import annotations

import pathlib
import sys
import tempfile
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

import activity_edit as ae  # noqa: E402
import edit_common as ec  # noqa: E402
import oeq_edit as oe  # noqa: E402
import word_edit as we  # noqa: E402
import word_ids as W  # noqa: E402
from report.plain import plain  # noqa: E402

CASES: list = []


def case(fn):
    CASES.append(fn)
    return fn


def sheet(rows: list[dict], *, columns: list[str] | None = None,
          split: list[str] | None = None) -> Path:
    """A minimal Vocab Changes workbook, written where load_rows can read it."""
    cols = columns or ["Book", "Book ID", "word_id", "current_word", "fixed_word",
                       "current_pos", "fixed_pos", "current_def", "fixed_def",
                       "current_kor", "fixed_kor", "current_vie", "fixed_vie",
                       "Error Types", "Details"]
    wb = Workbook()
    ws = wb.active
    ws.title = we.SHEET
    for c, h in enumerate(cols, 1):
        ws.cell(we.HEADER_ROW, c, h)
    for i, row in enumerate(rows):
        for c, h in enumerate(cols, 1):
            ws.cell(we.HEADER_ROW + 1 + i, c, row.get(h, ""))
    sp = wb.create_sheet(we.SPLIT_SHEET)
    for c, h in enumerate(["Book", "Book ID", "word_id"], 1):
        sp.cell(we.HEADER_ROW, c, h)
    for i, wid in enumerate(split or []):
        sp.cell(we.HEADER_ROW + 1 + i, 3, wid)
    path = Path(tempfile.mkdtemp()) / "report.xlsx"
    wb.save(path)
    return path


# --------------------------------------------------------------------------------------
# What the database will and will not accept
# --------------------------------------------------------------------------------------


@case
def key_is_case_sensitive():
    """A unique constraint in Postgres compares case. Lowercasing first made the check
    refuse seven edits the database would have taken."""
    assert we._key("bonnet(s)", "noun", "a hat") != we._key("Bonnet(s)", "noun", "a hat")
    assert we._key("bonnet(s)", "noun", "a hat") == we._key("bonnet(s)", "Noun", "a hat")


@case
def collision_ignores_the_row_itself():
    """An entry always matches its own key; only another id is a collision."""
    row = {"word_id": "10", "fixed_word": "", "current_word": "pea",
           "fixed_pos": "", "current_pos": "noun",
           "fixed_definition": "a vegetable", "current_definition": "pea"}
    key = we._key("pea", "noun", "a vegetable")
    assert we.collides(row, {key: ["10"]}) == []
    assert we.collides(row, {key: ["10", "99"]}) == ["99"]


# --------------------------------------------------------------------------------------
# Which rows a run may touch
# --------------------------------------------------------------------------------------


@case
def wrong_sense_is_held_back():
    path = sheet([{"word_id": "1", "current_def": "old", "fixed_def": "new",
                   "Error Types": "Wrong Sense"},
                  {"word_id": "2", "current_def": "old", "fixed_def": "new",
                   "Error Types": "Grammar"}])
    got = [r["word_id"] for r in we.load_rows(path, we.SHEET)]
    assert got == ["2"], got


@case
def cleared_ids_come_through_the_hold():
    wid = sorted(we.WRONG_SENSE_CLEARED)[0]
    path = sheet([{"word_id": wid, "current_def": "old", "fixed_def": "new",
                   "Error Types": "Wrong Sense"}])
    assert [r["word_id"] for r in we.load_rows(path, we.SHEET)] == [wid]


@case
def missing_error_types_column_is_refused():
    """The hold reads that column, and a lookup for an absent column returns empty. A
    report without it would once have written every entry the hold exists to keep back."""
    cols = ["word_id", "current_word", "fixed_word", "current_pos", "fixed_pos",
            "current_def", "fixed_def"]
    path = sheet([{"word_id": "1", "fixed_def": "new"}], columns=cols)
    try:
        we.load_rows(path, we.SHEET)
    except SystemExit as exc:
        assert "Error Types" in str(exc), exc
    else:
        raise AssertionError("a report with no Error Types column was accepted")


@case
def split_entries_are_never_edited_in_place():
    path = sheet([{"word_id": "7", "current_def": "old", "fixed_def": "new",
                   "Error Types": "Wrong Entry"}], split=["7A", "7B"])
    assert we.load_rows(path, we.SHEET) == []


@case
def rows_with_nothing_to_change_are_dropped():
    path = sheet([{"word_id": "1", "current_def": "same", "Error Types": "Grammar"}])
    assert we.load_rows(path, we.SHEET) == []


# --------------------------------------------------------------------------------------
# Resuming
# --------------------------------------------------------------------------------------


@case
def a_word_is_done_only_when_every_pending_field_is_written():
    """Treating a word as finished because it was saved once left three entries with an
    empty Vietnamese box that no later run would ever pick up."""
    row = {"word_id": "5", "fixed_word": "", "fixed_pos": "",
           "fixed_definition": "d", "fixed_kor": "", "fixed_vie": "v"}
    assert we.already_done(row, {"5": {"definition": "d", "vie": "v"}})
    assert not we.already_done(row, {"5": {"definition": "d"}})
    assert not we.already_done(row, {"5": {"definition": "other", "vie": "v"}})
    assert not we.already_done(row, {})


# --------------------------------------------------------------------------------------
# Refusing to write over an entry that has moved on
# --------------------------------------------------------------------------------------


@case
def identity_check_compares_what_the_report_recorded():
    row = {"word_id": "3", "current_word": "pea", "current_pos": "noun",
           "current_definition": "a vegetable", "current_kor": "", "current_vie": "",
           "fixed_word": "", "fixed_pos": "", "fixed_definition": "x",
           "fixed_kor": "", "fixed_vie": ""}
    same = {"word_id": "3", "word": "pea", "pos": "noun",
            "definition": "a vegetable", "kor": "", "vie": ""}
    assert we.check_matches(row, same) is None
    assert we.check_matches(row, {**same, "word_id": "4"})
    assert we.check_matches(row, {**same, "definition": "moved on"})


@case
def a_translation_is_only_checked_when_it_is_being_replaced():
    """The translations were read in a separate pass, so a Korean edited by someone else
    must not block an unrelated English fix."""
    base = {"word_id": "3", "current_word": "pea", "current_pos": "noun",
            "current_definition": "a vegetable", "current_kor": "A", "current_vie": "",
            "fixed_word": "", "fixed_pos": "", "fixed_definition": "x", "fixed_vie": ""}
    live = {"word_id": "3", "word": "pea", "pos": "noun",
            "definition": "a vegetable", "kor": "B", "vie": ""}
    assert we.check_matches({**base, "fixed_kor": ""}, live) is None
    assert we.check_matches({**base, "fixed_kor": "C"}, live)


# --------------------------------------------------------------------------------------
# Matching a vocabulary entry to its global id
# --------------------------------------------------------------------------------------


def index(*words):
    return W.build_index([{"id": i, "word": w, "pos": p, "definition": d}
                          for i, w, p, d in words])


@case
def exact_match_wins():
    tiers = index(("1", "root", "noun", "the part of a plant"))
    assert W.resolve("root", "the part of a plant", tiers)[0] == "1"


@case
def part_of_speech_breaks_a_tie():
    """The list holds 522 word-and-definition pairs twice, nearly all of them filed under
    two different parts of speech; the workbook's own tag settles which is meant."""
    tiers = index(("1", "root", "verb", "the part of a plant"),
                  ("2", "root", "noun", "the part of a plant"))
    assert W.resolve("root", "the part of a plant", tiers, "noun")[0] == "2"
    assert W.resolve("root", "the part of a plant", tiers)[0] is None


@case
def a_true_duplicate_resolves_to_nothing():
    """Same word, part of speech and definition twice over: an arbitrary id here is the
    confusion the ids exist to remove."""
    tiers = index(("1", "be afraid of", "phrase", "to feel fear"),
                  ("2", "be afraid of", "phrase", "to feel fear"))
    assert W.resolve("be afraid of", "to feel fear", tiers, "phrase")[0] is None


@case
def word_alone_never_picks_a_sense():
    """With the definition already failed to match, word plus part of speech would choose
    a meaning rather than confirm one."""
    tiers = index(("1", "bark", "noun", "a tree's covering"),
                  ("2", "bark", "noun", "a dog's sound"))
    assert W.resolve("bark", "something else entirely", tiers, "noun")[0] is None


# --------------------------------------------------------------------------------------
# Rewriting finding text into plain language
# --------------------------------------------------------------------------------------


@case
def quoted_text_is_left_exactly_as_written():
    """Rewriting inside quotation marks corrupts the very text a fixer must search for:
    "a large circular structure" once became "a large explains itself in a circle"."""
    out = plain('above level: "a large circular structure"')
    assert '"a large circular structure"' in out, out


# --------------------------------------------------------------------------------------
# Staying consistent across a re-scrape
# --------------------------------------------------------------------------------------


@case
def an_unresolved_slot_does_not_keep_the_previous_word_s_id():
    """A re-scrape can put a different word in a slot. An id left over from the previous
    occupant is worse than none, because everything downstream trusts the column."""
    from openpyxl import Workbook

    import rs_scrape as rs

    wb = Workbook()
    ws = wb.active
    ws.title = "Vocab"
    for c, h in enumerate(["id", "W1", "POS1", "DEF1", "SENT1", "WID1"], 1):
        ws.cell(1, c, h)
    ws.cell(2, 1, "900")
    ws.cell(2, 2, "zebra")
    ws.cell(2, 3, "noun")
    ws.cell(2, 4, "a definition no global entry has")
    ws.cell(2, 6, 111111)
    headers = rs.header_map(ws)
    tiers = index(("222", "other", "noun", "x"))
    for r, slot, word, pos, dfn in W.vocab_entries(ws, headers):
        wid, _, _ = W.resolve(word, dfn, tiers, pos)
        if wid:
            ws.cell(r, headers[f"wid{slot}"], int(wid))
        else:
            ws.cell(r, headers[f"wid{slot}"]).value = None
    assert ws.cell(2, 6).value is None, ws.cell(2, 6).value


@case
def every_module_reads_the_same_number_of_word_slots():
    """A slot bound written out by hand in one module and imported in the others drifts:
    words past the hardcoded limit go missing from the reports without a word said."""
    import re

    import rs_scrape as rs

    source = pathlib.Path("make_reports.py").read_text(encoding="utf-8")
    hardcoded = re.findall(r"for s in range\(1, (\d+)\)", source)
    assert not hardcoded, f"make_reports.py hardcodes a slot count: {hardcoded}"
    assert rs.MAX_SLOTS >= 25


# --------------------------------------------------------------------------------------
# Rewriting the open-ended questions
# --------------------------------------------------------------------------------------


@case
def an_instruction_is_not_a_replacement_question():
    """Three findings carry the reviewer's note where the new question belongs. Typed
    out, that sentence becomes what the child is asked."""
    assert not oe.usable("Replace it with a question a student can answer from their "
                         "own experience before they open the book.")
    assert not oe.usable("Rewrite this to be simpler.")
    assert oe.usable("What do you think this story will be about?")
    assert oe.usable("Have you seen a desert? What did it look like?")


@case
def a_contested_question_is_left_alone_unless_it_was_settled_by_hand():
    """Two findings wanting different text is a decision, not something to resolve by
    whichever happened to be read first."""
    f = lambda bid, tgt, fix, typ: {
        "sheet": "OEC", "prefix": "Q", "book_id": bid, "book": "b", "target": tgt,
        "n": int(tgt[1:]), "type": typ, "fix": fix, "current": "old text"}

    def load(findings):
        import make_reports
        real = make_reports.load_findings
        make_reports.load_findings = lambda path, titles: findings
        try:
            return oe.load_edits([pathlib.Path("tests/test_logic.py")], {})
        finally:
            make_reports.load_findings = real

    books, notes = load([f("9", "Q1", "A one?", "Grammar"),
                         f("9", "Q1", "A two?", "Too Hard")])
    assert books == {}, books
    assert any("CHOSEN" in n for n in notes), notes

    # Agreeing findings are one edit, not a conflict.
    books, _ = load([f("9", "Q1", "Same?", "Grammar"), f("9", "Q1", "Same?", "Unclear")])
    assert books["9"]["questions"]["Q1"]["fix"] == "Same?"
    assert books["9"]["questions"]["Q1"]["types"] == ["Grammar", "Unclear"]

    # The instruction drops out, leaving the real rewrite uncontested.
    books, _ = load([f("9", "Q1", "A real one?", "Grammar"),
                     f("9", "Q1", "Replace it with something better.", "Requires Reading")])
    assert books["9"]["questions"]["Q1"]["fix"] == "A real one?"


@case
def a_settled_question_takes_the_reviewed_text():
    bid, tgt = next(iter(oe.CHOSEN))
    assert oe.CHOSEN[(bid, tgt)].endswith("?")


@case
def a_book_saved_for_one_question_still_owes_the_other():
    """The fault that once left three Vietnamese boxes empty for good: treating a book
    as finished because it was saved once."""
    book = {"book_id": "9", "book": "b", "questions": {"Q1": {}, "Q2": {}}}
    assert set(oe.outstanding(book, {"9": {"Q1"}})["questions"]) == {"Q2"}
    assert oe.outstanding(book, {"9": {"Q1", "Q2"}})["questions"] == {}
    assert set(oe.outstanding(book, {})["questions"]) == {"Q1", "Q2"}


@case
def a_spacing_fix_is_compared_exactly():
    """Several fixes correct spacing and nothing else. Collapsing whitespace before
    comparing would call such an edit verified without it ever being made."""
    assert not ec.same("what did you?  If no", "what did you? If no")
    assert ec.same(" trailing ", "trailing")


@case
def the_audio_remove_is_never_found_by_position():
    """Each question has two Remove buttons and the one beside the heading deletes the
    question. A question whose audio is already gone has only that one left."""
    src = pathlib.Path("oeq_edit.py").read_text(encoding="utf-8")
    body = src.split("FIND_JS")[1].split("COUNT_JS")[0]
    assert "change.parentElement" in body, "the audio Remove must be found via Change Audio"
    assert "removeButton: remove ? buttons.indexOf(remove) : -1" in body
    # and nothing may fall back to picking a Remove by index
    assert "'Remove')[1]" not in body and '"Remove")[1]' not in body


@case
def a_refused_save_is_read_out_of_the_body():
    """The endpoint answers HTTP 200 and puts its objection in the body, so a refusal
    and a success look identical from outside. Only the body tells them apart."""
    class Resp:
        url = "https://x/graphql-prod"
        def __init__(self, payload, method="POST"):
            self._p = payload
            self.request = type("R", (), {"method": method})()
        def json(self):
            return self._p

    class Page:
        def on(self, event, fn):
            self.fn = fn

    page = Page()
    m = ec.Mutation(page, "OpenEndedQuestions")
    page.fn(Resp({"errors": [{"message": "ELEVENLABS_API_KEY is not set",
                              "path": ["updateOpenEndedQuestionsContent"]}],
                  "data": None}))
    assert m.errors == ["ELEVENLABS_API_KEY is not set"]
    assert m.take() == ["ELEVENLABS_API_KEY is not set"]
    assert m.take() == [], "taking twice must not repeat the same refusal"
    # A success, and an error belonging to some other mutation, are both ignored.
    page.fn(Resp({"data": {"updateOpenEndedQuestionsContent": {"id": 1}}}))
    page.fn(Resp({"errors": [{"message": "nope", "path": ["wordListByBook"]}]}))
    assert m.take() == []


# --------------------------------------------------------------------------------------
# Context clue and text multiple choice
# --------------------------------------------------------------------------------------


@case
def a_field_name_maps_to_the_right_box():
    """AnsC4 is the third answer of question 4, and Q4 is that question's own text.
    Getting the letter-to-slot step wrong rewrites a different answer."""
    assert ae.parse_target("Q4") == (4, 0)
    assert ae.parse_target("AnsA4") == (4, 1)
    assert ae.parse_target("AnsC4") == (4, 3)
    assert ae.parse_target("AnsD10") == (10, 4)
    # AnsAll is expanded before it gets here, and nothing else is a field.
    assert ae.parse_target("AnsAll4") is None
    assert ae.parse_target("Main_Character") is None
    assert ae.parse_target("A9") is None


@case
def all_four_options_split_whether_or_not_they_are_on_separate_lines():
    """Four of the fifteen AnsAll cells put the options on one line. Splitting on
    newlines alone yields a single option and pushes all four into answer A."""
    lines = "\n".join(["A. one", "B. two", "C. three", "D. four"])
    flat = "A. one B. two C. three D. four"
    for text in (lines, flat):
        got = ae.split_options(text)
        assert got == {"A": "one", "B": "two", "C": "three", "D": "four"}, got
    # A cell that is prose rather than options yields nothing, which is what makes the
    # book-2841 finding fall out rather than be typed in.
    assert set(ae.split_options("Replace the question and all four options.")) != set("ABCD")


@case
def an_ansall_finding_only_rewrites_the_options_that_changed():
    """A punctuation fix on two options must not rewrite the other two."""
    import make_reports

    finding = {"sheet": "TMC", "prefix": "AnsAll", "book_id": "9", "book": "b",
               "target": "AnsAll8", "n": 8, "type": "Punctuation",
               "current": "\n".join(["A. one.", "B. two.", "C. three", "D. four"]),
               "fix": "\n".join(["A. one", "B. two", "C. three", "D. four"])}
    real = make_reports.load_findings
    make_reports.load_findings = lambda path, titles: [finding]
    try:
        books, _ = ae.load_edits(ae.ACTIVITIES["TMC"],
                                 [pathlib.Path("tests/test_logic.py")], {})
    finally:
        make_reports.load_findings = real
    fields = books["9"]["fields"]
    assert set(fields) == {"AnsA8", "AnsB8"}, set(fields)
    assert fields["AnsA8"]["fix"] == "one"
    assert fields["AnsA8"]["slot"] == 1
    assert fields["AnsB8"]["slot"] == 2


@case
def a_finding_with_no_suggested_fix_is_never_written():
    """Thirteen context-clue findings say what is wrong without saying what to put
    there. Writing an empty box would erase the item."""
    import make_reports

    finding = {"sheet": "CC", "prefix": "Q", "book_id": "9", "book": "b",
               "target": "Q1", "n": 1, "type": "Grammar", "current": "text",
               "fix": "   ", "details": "grammar problem"}
    real = make_reports.load_findings
    make_reports.load_findings = lambda path, titles: [finding]
    try:
        books, notes = ae.load_edits(ae.ACTIVITIES["CC"],
                                     [pathlib.Path("tests/test_logic.py")], {})
    finally:
        make_reports.load_findings = real
    # Not "no books at all": the amendments are injected on every load, by design.
    assert "9" not in books, books
    assert any("no suggested fix" in n for n in notes), notes


@case
def an_amendment_outlives_the_journal_that_says_it_is_done():
    """An amendment exists because what went in was wrong, and the journal's whole
    record of that field is that it was written. Filtering on the journal would drop
    every amendment before it ever ran."""
    book = {"book_id": "9", "book": "b",
            "fields": {"Q1": {"amended": True}, "Q2": {}}}
    kept = ae.outstanding(book, {"9": {"Q1", "Q2"}})["fields"]
    assert set(kept) == {"Q1"}, kept


@case
def an_amendment_says_what_went_in_not_what_the_report_quoted():
    """The page holds what this script last wrote there, so that is what the identity
    check has to expect -- the report's original quote is two edits stale."""
    for (key, bid, target), (was, should) in ae.AMENDED.items():
        assert key in ae.ACTIVITIES, key
        assert ae.parse_target(target) is not None, key
        assert was != should, key
    import make_reports

    real = make_reports.load_findings
    make_reports.load_findings = lambda path, titles: []
    try:
        books, _ = ae.load_edits(ae.ACTIVITIES["CC"],
                                 [pathlib.Path("tests/test_logic.py")], {})
    finally:
        make_reports.load_findings = real
    for (key, bid, target), (was, should) in ae.AMENDED.items():
        if key != "CC":
            continue
        got = books[bid]["fields"][target]
        assert got["current"] == was and got["fix"] == should, got


@case
def the_hand_written_decisions_are_reachable():
    """CHOSEN is keyed by activity as well as book and field; keying it by book alone
    would let a context-clue decision answer for a multiple-choice one."""
    for key in ae.CHOSEN:
        assert key[0] in ae.ACTIVITIES, key
        assert ae.parse_target(key[2]) is not None, key
    for key in ae.UNUSABLE:
        assert key[0] in ae.ACTIVITIES, key
    for key in ae.RESOLVED:
        assert key[0] in ae.ACTIVITIES, key
        # A9 is an answer pill rather than a field, and pills are not something this
        # script edits; it is recorded here because it was outstanding and is not now.
        assert ae.parse_target(key[2]) is not None or key[2].startswith("A"), key
    for key in ae.DECLINED:
        assert key[0] in ae.ACTIVITIES, key


@case
def a_book_saved_for_one_field_still_owes_the_others():
    book = {"book_id": "9", "book": "b", "fields": {"Q1": {}, "AnsC2": {}}}
    assert set(ae.outstanding(book, {"9": {"Q1"}})["fields"]) == {"AnsC2"}
    assert ae.outstanding(book, {"9": {"Q1", "AnsC2"}})["fields"] == {}


if __name__ == "__main__":
    failed = []
    for fn in CASES:
        try:
            fn()
            print(f"  ok   {fn.__name__}")
        except Exception:                                          # noqa: BLE001
            failed.append(fn.__name__)
            print(f"  FAIL {fn.__name__}")
            traceback.print_exc()
    print(f"\n{len(CASES) - len(failed)}/{len(CASES)} passed")
    sys.exit(1 if failed else 0)

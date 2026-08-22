"""Parse the hand-written QA notes in Level_2_tracking.xlsx into structured findings.

The tracking workbook is a human working document, not an export, and the encoding drifts:
roughly 84% of the Listen & Read entries match the documented shape and the rest carry
typos, missing fields, or a second shape entirely. Everything this module knows about that
drift is written down here rather than discovered again each time.

Two conventions matter more than the grammar:

  * a cell containing exactly ``o`` means "checked, found nothing" and ``?`` means
    "unresolved". Treating non-empty as an error turns ~40 real findings into 87 rows of
    noise, so both are dropped.
  * a newline inside a cell is usually a soft wrap in a quoted sentence, not an entry
    separator. Entries are split on a page-header lookahead instead.

    uv run python manual_qa.py                 # summary of what parses
    uv run python manual_qa.py --dump LR       # every parsed LR finding
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

TRACKING = Path("references/Level_2_tracking.xlsx")
BOOKS_SHEET = "Books"
HEADER_ROW = 4
ID_COL, TITLE_COL, LEVEL_COL = 1, 2, 4      # column C is a second, empty "Title"

# Cells that record a status rather than a finding.
CLEAN, UNRESOLVED = "o", "?"

# An entry starts at a page header: P12, p5*, P4#, Pall, or a bare P with no number.
ENTRY_START = re.compile(r"(?m)^(?=[Pp](?:all|\d+)?[*#]?[_:])")
HEADER = re.compile(r"^[Pp](?P<page>all|\d*)(?P<mark>[*#]?)(?:_(?P<type>[A-Za-z][\w-]*))?:?\s*")

# Nine surface tokens for five real Listen & Read faults.
LR_TYPES = {
    "cutoff-minor": "Minor TTS Cutoff",
    "cutoff_minor": "Minor TTS Cutoff",
    "cutoff-minior": "Minor TTS Cutoff",
    "cutoff-severe": "Severe TTS Cutoff",
    "cutoff_severe": "Severe TTS Cutoff",
    "cutoff": "Severe TTS Cutoff",          # bare: the documented default
    "start-cutoff": "Start TTS Cutoff",
    "distortion": "TTS Pronunciation",
    "distorted": "TTS Pronunciation",
    "punct": "Text",
    "whole-page-highlight": "Other",
    "text-tts-mismatch": "Text-TTS Mismatch",
    "quality_moderate": "Other",
}
DEFAULT_LR_TYPE = "Severe TTS Cutoff"       # stated rule: a missing type means severe

# `text-tts-mismatch: text="...", TTS="..."` carries no page number at all.
MISMATCH = re.compile(r'^\s*text-tts-mismatch:\s*text=(?P<text>.*?),\s*TTS=(?P<tts>.*)$',
                      re.S | re.I)
# `P7: Galaxies -> galaxies [capitalization]`
FIX_TAIL = re.compile(r"^(?P<text>.*?)\s*->\s*(?P<fix>.*?)\s*(?:\[(?P<tag>[^\]]+)\])?\s*$", re.S)
# An escaped control character that leaked into the export, not book content.
JUNK = re.compile(r"_x[0-9a-fA-F]{4}_")


@dataclass
class Finding:
    book_id: str
    book: str
    activity: str
    page: str = ""
    type: str = "Other"
    text: str = ""
    fix: str = ""
    details: str = ""
    source_column: str = ""


@dataclass
class ParseReport:
    findings: list[Finding] = field(default_factory=list)
    skipped_clean: int = 0
    skipped_unresolved: int = 0
    unparsed: list[tuple[str, str]] = field(default_factory=list)


def _clean(value) -> str:
    if value is None:
        return ""
    return JUNK.sub("", str(value)).strip()


def split_entries(cell: str) -> list[str]:
    """One entry per page header, keeping soft-wrapped sentences whole."""
    parts = [p.strip() for p in ENTRY_START.split(cell) if p.strip()]
    return parts or ([cell.strip()] if cell.strip() else [])


def normalise_type(token: str) -> str:
    if not token:
        return DEFAULT_LR_TYPE
    key = token.strip().lower()
    if key in LR_TYPES:
        return LR_TYPES[key]
    # cutoff_minor / cutoff-minior style drift
    for surface, canonical in LR_TYPES.items():
        if key.replace("_", "-") == surface.replace("_", "-"):
            return canonical
    return ""


def parse_lr_entry(entry: str, activity: str, column: str, book_id: str, book: str,
                   default_type: str = "", default_details: str = ""):
    """One Listen & Read entry -> a Finding, or None if it cannot be read.

    `default_type` carries the meaning of the column itself. The LR columns encode the
    fault in each entry, but the LRA ones encode it in the column heading and most of
    their entries are a bare quoted sentence with no page header at all.
    """
    mismatch = MISMATCH.match(entry)
    if mismatch:
        text = mismatch.group("text").strip().strip('"')
        tts = mismatch.group("tts").strip().strip('"')
        return Finding(book_id, book, activity, "", "Text-TTS Mismatch", text, tts,
                       "on-page text and the audio do not match", column)

    header = HEADER.match(entry)
    if not header:
        if not default_type:
            return None
        return Finding(book_id, book, activity, "", default_type,
                       entry.strip().strip('"'), "", default_details, column)
    page = header.group("page") or ""
    if entry[:4].lower().startswith("pall"):
        page = "all"
    body = entry[header.end():].strip()
    token = header.group("type") or ""
    kind = normalise_type(token)
    details = ""
    if not kind:
        # The type slot held sentence text, not a type -- keep it and fall back.
        body = f"{token} {body}".strip()
        kind = DEFAULT_LR_TYPE
        details = "error type missing in the source note"
    elif not token:
        details = "error type missing in the source note"

    fix = ""
    if column.endswith("text"):
        kind = "Text"
        tail = FIX_TAIL.match(body)
        if tail and tail.group("fix"):
            body, fix = tail.group("text").strip(), tail.group("fix").strip()
            if tail.group("tag"):
                details = tail.group("tag").strip()
    return Finding(book_id, book, activity, page, kind,
                   body.strip().strip('"'), fix, details, column)


def parse_column(rows, header_index: dict[str, int], column: str, activity: str,
                 report: ParseReport, default_type: str = "",
                 default_details: str = "") -> None:
    col = header_index[column]
    for book_id, book, values in rows:
        cell = _clean(values.get(col))
        if not cell:
            continue
        if cell == CLEAN:
            report.skipped_clean += 1
            continue
        if cell == UNRESOLVED:
            report.skipped_unresolved += 1
            continue
        for entry in split_entries(cell):
            found = parse_lr_entry(entry, activity, column, book_id, book,
                                   default_type, default_details)
            if found is None:
                report.unparsed.append((f"{book_id} {book} [{column}]", entry))
            else:
                report.findings.append(found)


def load_books(path: Path = TRACKING):
    """(header_index, rows) where rows is [(book_id, title, {col: value})]."""
    ws = load_workbook(path, data_only=True)[BOOKS_SHEET]
    header_index = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(HEADER_ROW, c).value
        if name:
            header_index.setdefault(str(name).strip(), c)
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        raw = ws.cell(r, ID_COL).value
        if raw is None:
            continue
        book_id = str(int(raw)) if isinstance(raw, float) and raw.is_integer() else str(raw).strip()
        values = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rows.append((book_id, _clean(ws.cell(r, TITLE_COL).value), values))
    return header_index, rows


# (column, activity, type the column itself implies, details for that type)
LR_COLUMNS = [
    ("LR - TTS", "LR", "", ""),
    ("LR - text", "LR", "Text", ""),
    ("LR - other", "LR", "Other", ""),
]
# The LRA columns name the fault in their heading, and most entries are a bare quoted
# sentence, so the type comes from the column rather than from the entry.
LRA_COLUMNS = [
    ("LRA - TTS too long", "LRA", "Other",
     "the audio keeps playing after the highlighted text ends"),
    ("LRA - TTS cut off", "LRA", "Severe TTS Cutoff",
     "the audio stops before the sentence finishes"),
    ("LRA - text", "LRA", "Text", ""),
    ("LRA - other", "LRA", "Other", ""),
]


def parse_listen_and_read(path: Path = TRACKING) -> tuple[ParseReport, ParseReport]:
    header_index, rows = load_books(path)
    lr, lra = ParseReport(), ParseReport()
    for column, activity, kind, details in LR_COLUMNS:
        parse_column(rows, header_index, column, activity, lr, kind, details)
    for column, activity, kind, details in LRA_COLUMNS:
        parse_column(rows, header_index, column, activity, lra, kind, details)
    return lr, lra


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--dump", choices=["LR", "LRA"], help="print every parsed finding")
    args = parser.parse_args()

    lr, lra = parse_listen_and_read()
    for name, rep in (("LR", lr), ("LRA", lra)):
        kinds = {}
        for f in rep.findings:
            kinds[f.type] = kinds.get(f.type, 0) + 1
        print(f"\n=== {name}: {len(rep.findings)} finding(s) across "
              f"{len({f.book_id for f in rep.findings})} book(s)")
        print(f"    skipped: {rep.skipped_clean} clean, {rep.skipped_unresolved} unresolved")
        for k, n in sorted(kinds.items(), key=lambda x: -x[1]):
            print(f"    {n:4}  {k}")
        if rep.unparsed:
            print(f"    !! {len(rep.unparsed)} unparsed:")
            for where, entry in rep.unparsed:
                print(f"       {where}: {entry[:90]!r}")

    if args.dump:
        rep = lr if args.dump == "LR" else lra
        print(f"\n=== every {args.dump} finding ===")
        for f in rep.findings:
            print(f"  {f.book_id:>5} {f.book[:26]:26} P{f.page or '-':4} {f.type:20} "
                  f"{f.text[:60]!r}")


if __name__ == "__main__":
    main()


# --------------------------------------------------------------------------------------
# Per-activity notes.
#
# Each column below uses its own shorthand, and none of them is quite the shape the
# others use. The parsers are deliberately small and literal rather than one clever
# grammar, because the cost of guessing wrong is a finding pointed at the wrong word.
# --------------------------------------------------------------------------------------

# `too-hard: museum`, `word-in-def: northern`, `crop: "scale"`, `redo: league`
TAGGED = re.compile(r'^\s*(?P<tag>[a-z][\w-]*)\s*:\s*(?P<value>.*)$', re.I | re.S)
# `browse, browsed browsing: browse, browsed, browsing [add missing comma]`
REPLACEMENT = re.compile(r'^(?P<bad>.*?):\s*(?P<good>.*?)\s*(?:\[(?P<tag>[^\]]+)\])?\s*$', re.S)
# `Q8: ... [punctuation]` / `Q3_B: bad -> good [grammar]` / `Q3_multi: ...` / `Q_ALL: ...`
TMC_ENTRY = re.compile(
    r'^\s*Q(?P<slot>\d+|_ALL)(?:_(?P<choice>[A-D]|multi|ALL))?\s*:\s*(?P<body>.*)$', re.S)
# A vocabulary word carrying an inflection: "wish(es)", "snatch, (snatching)".
INFLECTED = re.compile(r'\([^)]*\)\s*$|,\s*\(?[^,()]+\)?\s*$')

VOCAB_TAGS = {
    "too-hard": "Too Hard",
    "word-in-def": "Answer Given",
    "wrong": "Other",
    "wrong-sense": "Wrong Sense",
    "cut-off-word": "Part of Speech",
    "similar-meanings": "Lack of Context",
}


def _strip_quotes(text: str) -> str:
    return text.strip().strip('"').strip("\u201c\u201d").strip()


def _entries(cell: str) -> list[str]:
    """Split a per-activity cell into entries.

    These columns really are newline-separated, unlike the Listen & Read ones -- except
    where a `->` replacement wrapped onto the next line, which would otherwise split one
    note into an entry with no fix and an orphan with no question number.
    """
    out: list[str] = []
    for line in cell.splitlines():
        line = line.strip().rstrip(";").strip()
        if not line:
            continue
        if out and out[-1].endswith("->"):
            out[-1] = f"{out[-1]} {line}"
        else:
            out.append(line)
    return out


def is_inflected(word: str) -> bool:
    """True for words being re-recorded anyway, e.g. wish(es), stake a claim, (staked...).

    Deliberately narrow: `wish(es)` carries a genuine trailing note in the source
    (`(pronounced "wish ess")`) and a blanket "any parenthesis" test would also strip
    multi-word entries like `ghost town` that must survive.
    """
    return bool(INFLECTED.search(word.strip()))


@dataclass
class ActivityFinding:
    """A manual note aimed at a row of one of the scraped activity sheets."""
    book_id: str
    book: str
    sheet: str            # OEC / CC / Vocab / TMC
    word: str = ""        # the vocabulary word, where the note names one
    slot_hint: str = ""   # e.g. "Q8", "AnsB3" -- resolved to a real column later
    type: str = "Other"
    text: str = ""        # the offending text as the note quotes it
    fix: str = ""
    details: str = ""
    source_column: str = ""


def parse_vocab_tts(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`Vocab - TTS`: words whose audio needs re-recording."""
    out, col = [], header_index["Vocab - TTS"]
    for book_id, book, values in rows:
        cell = _clean(values.get(col))
        if not cell or cell in (CLEAN, UNRESOLVED):
            report.skipped_clean += cell == CLEAN
            report.skipped_unresolved += cell == UNRESOLVED
            continue
        for entry in _entries(cell):
            m = TAGGED.match(entry)
            if not m or m.group("tag").lower() != "redo":
                report.unparsed.append((f"{book_id} {book} [Vocab - TTS]", entry))
                continue
            value = m.group("value")
            note = ""
            # `"wish(es)" (pronounced "wish ess")` -- keep the note, drop it from the word
            trailing = re.search(r'^(?P<word>.*?[")])\s*\((?P<note>[^()]*(?:\([^()]*\))?[^()]*)\)\s*$',
                                 value)
            if trailing:
                value, note = trailing.group("word"), trailing.group("note").strip()
            word = _strip_quotes(value)
            if is_inflected(word):
                continue          # being re-recorded anyway when the inflection is fixed
            details = "the recorded audio for this word needs redoing"
            if note:
                details += f" -- {note}"
            out.append(ActivityFinding(book_id, book, "Vocab", word=word, type="Other",
                                       text=word, details=details,
                                       source_column="Vocab - TTS"))
    return out


def parse_vocab_images(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`Vocab - Images`: pictures needing a re-crop or a replacement."""
    verbs = {"crop": "the picture for this word needs cropping",
             "redo": "the picture for this word needs replacing"}
    out, col = [], header_index["Vocab - Images"]
    for book_id, book, values in rows:
        cell = _clean(values.get(col))
        if not cell or cell in (CLEAN, UNRESOLVED):
            report.skipped_clean += cell == CLEAN
            report.skipped_unresolved += cell == UNRESOLVED
            continue
        for entry in _entries(cell):
            m = TAGGED.match(entry)
            if not m or m.group("tag").lower() not in verbs:
                report.unparsed.append((f"{book_id} {book} [Vocab - Images]", entry))
                continue
            word = _strip_quotes(m.group("value"))
            out.append(ActivityFinding(book_id, book, "Vocab", word=word, type="Other",
                                       text=word, details=verbs[m.group("tag").lower()],
                                       source_column="Vocab - Images"))
    return out


def parse_vocab_notes(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`Vocab - Definition` (AI), `Vocab - Text` (AH), `WMM - cut-off word` (AC), `WMM - other` (AD).

    AC and AD are not cleanly separated in the source -- book 929 records two
    `cut-off-word:` notes in AD that belong in AC -- so both are read the same way.
    """
    out = []
    # AI: `too-hard: museum` -- a tag plus the word it applies to.
    for column in ("Vocab - Definition",):
        col = header_index[column]
        for book_id, book, values in rows:
            cell = _clean(values.get(col))
            if not cell or cell in (CLEAN, UNRESOLVED):
                continue
            for entry in _entries(cell):
                m = TAGGED.match(entry)
                if not m:
                    report.unparsed.append((f"{book_id} {book} [{column}]", entry))
                    continue
                tag = m.group("tag").lower()
                word = _strip_quotes(m.group("value"))
                out.append(ActivityFinding(
                    book_id, book, "Vocab", word=word,
                    type=VOCAB_TAGS.get(tag, "Other"), text=word,
                    details=f"manual QA note: {tag}", source_column=column))

    # AH: `aike: alike [spelling]` -- the word text itself is malformed.
    col = header_index["Vocab - Text"]
    for book_id, book, values in rows:
        cell = _clean(values.get(col))
        if not cell or cell in (CLEAN, UNRESOLVED):
            continue
        for entry in _entries(cell):
            m = REPLACEMENT.match(entry)
            if not m:
                report.unparsed.append((f"{book_id} {book} [Vocab - Text]", entry))
                continue
            bad, good = m.group("bad").strip(), m.group("good").strip()
            tag = (m.group("tag") or "").strip()
            out.append(ActivityFinding(
                book_id, book, "Vocab", word=bad, type="Part of Speech", text=bad,
                fix=good, details=tag or "the word text is malformed",
                source_column="Vocab - Text"))

    # AC / AD: bare trailing-comma words, and `tag: value` notes.
    for column in ("WMM - cut-off word", "WMM - other"):
        col = header_index[column]
        for book_id, book, values in rows:
            cell = _clean(values.get(col))
            if not cell or cell in (CLEAN, UNRESOLVED):
                continue
            for entry in _entries(cell):
                m = TAGGED.match(entry)
                if m and m.group("tag").lower() in VOCAB_TAGS:
                    tag = m.group("tag").lower()
                    word = _strip_quotes(m.group("value"))
                    out.append(ActivityFinding(
                        book_id, book, "Vocab", word=word,
                        type=VOCAB_TAGS[tag], text=word,
                        details=f"manual QA note: {tag}", source_column=column))
                else:
                    word = _strip_quotes(entry)
                    out.append(ActivityFinding(
                        book_id, book, "Vocab", word=word.rstrip(","),
                        type="Part of Speech", text=word,
                        fix=word.rstrip(", "),
                        details="the word tile ends in a stray comma, from a dropped "
                                "inflection",
                        source_column=column))
    return out


TMC_REASONS = {
    "punctuation": "Punctuation", "grammar": "Grammar", "capitalization": "Capitalization",
    "spelling": "Other", "context missing": "Unclear",   # TMC has no "Requires Reading" type
    "multiple_correct": "Unclear", "inconsistent punctuation": "Punctuation",
    "grammar/punctuation": "Grammar",
}


def parse_tmc_notes(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`TMC - Question` (AE) and `TMC - Answer` (AF)."""
    out = []
    for column, default_target in (("TMC - Question", "Q"), ("TMC - Answer", "Ans")):
        col = header_index[column]
        for book_id, book, values in rows:
            cell = _clean(values.get(col))
            if not cell or cell in (CLEAN, UNRESOLVED):
                continue
            for entry in _entries(cell):
                m = TMC_ENTRY.match(entry)
                if not m:
                    report.unparsed.append((f"{book_id} {book} [{column}]", entry))
                    continue
                slot, choice, body = m.group("slot"), m.group("choice"), m.group("body")
                if slot == "_ALL" or choice in ("ALL", "multi"):
                    # A whole-question note: no single cell to point at.
                    hint = f"Q{slot}" if slot != "_ALL" else ""
                else:
                    hint = f"Ans{choice}{slot}" if choice else f"Q{slot}"

                tag = ""
                tail = re.search(r"[\[(]([^\])]+)[\])]\s*$", body)
                if tail:
                    tag = tail.group(1).strip().lower()
                    body = body[: tail.start()].strip()
                text, fix = body, ""
                if "->" in body:
                    text, fix = (p.strip() for p in body.split("->", 1))
                out.append(ActivityFinding(
                    book_id, book, "TMC", slot_hint=hint,
                    type=TMC_REASONS.get(tag, "Other"),
                    text=text.strip(), fix=" ".join(fix.split()),
                    details=tag or "manual QA note", source_column=column))
    return out


def parse_cc_notes(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`CC - Sentence Grammar/punct` (Z) and `CC - Word in Sentence` (AA).

    Both use `answerword: sentence with ___ in it`, so the answer word identifies which
    A[#] slot the note is about and the sentence is the Q[#] beside it.
    """
    out = []
    for column, kind in (("CC - Sentence Grammar/punct", "Grammar"),
                         ("CC - Word in Sentence", "Answer Given")):
        col = header_index[column]
        for book_id, book, values in rows:
            cell = _clean(values.get(col))
            if not cell or cell in (CLEAN, UNRESOLVED):
                continue
            for entry in _entries(cell):
                # Nearly always `word: sentence`, but one book uses `word - sentence`.
                answer, _, sentence = entry.partition(":")
                if not sentence.strip() and " - " in entry:
                    answer, _, sentence = entry.partition(" - ")
                if not sentence.strip():
                    report.unparsed.append((f"{book_id} {book} [{column}]", entry))
                    continue
                out.append(ActivityFinding(
                    book_id, book, "CC", word=answer.strip(), type=kind,
                    text=sentence.strip(),
                    details=("the answer word or its root also appears in the sentence"
                             if kind == "Answer Given"
                             else "grammar or punctuation problem in the sentence"),
                    source_column=column))
    return out


def parse_oec_notes(rows, header_index, report: ParseReport) -> list[ActivityFinding]:
    """`OEC Story Q` (I): which warm-up questions need the story to have been read.

    Open-Ended Questions runs *before* reading, so a question that cannot be answered
    without the book is a fault, and the reviewer wrote down the question numbers where
    that happens. `None` is the opposite: they looked and found none, which is a clean
    result and must not become a finding.
    """
    out, col = [], header_index["OEC Story Q"]
    for book_id, book, values in rows:
        cell = _clean(values.get(col))
        if not cell or cell in (CLEAN, UNRESOLVED) or cell.casefold() == "none":
            report.skipped_clean += cell.casefold() == "none"
            continue
        slots = re.findall(r"[Qq]\s*(\d+)", cell)
        if not slots:
            report.unparsed.append(("OEC Story Q", f"{book_id}: {cell}"))
            continue
        for n in slots:
            out.append(ActivityFinding(
                book_id, book, "OEC", slot_hint=f"Q{n}", type="Requires Reading",
                details="this warm-up question cannot be answered without having read "
                        "the book, but the activity comes before reading",
                fix="Replace it with a question a student can answer from their own "
                    "experience before they open the book.",
                source_column="OEC Story Q"))
    return out


def parse_activities(path: Path = TRACKING):
    """Every per-activity manual note, plus a parse report."""
    header_index, rows = load_books(path)
    report = ParseReport()
    out = []
    out += parse_vocab_tts(rows, header_index, report)
    out += parse_vocab_images(rows, header_index, report)
    out += parse_vocab_notes(rows, header_index, report)
    out += parse_tmc_notes(rows, header_index, report)
    out += parse_cc_notes(rows, header_index, report)
    out += parse_oec_notes(rows, header_index, report)
    return out, report

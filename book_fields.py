"""Scrape the student-visible book fields into the OEC sheet.

The OEC activity shows the student two things the activity-edit page never carries:
the book's main-character list and its preview text. Both live on the book record at
/books/<id>/edit, so they cannot be reached by rs_scrape.py, whose every URL is
/activities/<id>/<segment>/edit. This script fills that gap and writes two flat
columns -- Main_Character and Prev_Text -- beside each OEC row.

Everything about pacing, login, session reuse and workbook safety is imported from
rs_scrape rather than reimplemented, so both scripts hit the server the same way.

One deliberate difference from rs_scrape: this script will CREATE its two columns if
the sheet lacks them, inserting them directly after `id`. rs_scrape never restructures
a workbook, but these columns are new to an established layout and every existing
workbook predates them. Pass --no-add-columns to require them up front instead.

    uv run python book_fields.py --dry-run          # report, write nothing
    uv run python book_fields.py                    # every id in the OEC sheet
    uv run python book_fields.py --books 697,1204   # just these
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright

import rs_scrape as rs

# --------------------------------------------------------------------------------------
# Configuration.
# --------------------------------------------------------------------------------------

SHEET = "OEC"

# Placeholder text identifying each field, and the column its value goes in. Matched
# with get_by_placeholder so the tag does not matter: main character is an <input> and
# preview text is a <textarea>, and either could change without breaking this.
FIELDS: tuple[tuple[str, str], ...] = (
    (
        "Main_Character",
        "Enter character names separated by commas. E.g., Charlie, Alice, Cat.",
    ),
    ("Prev_Text", "Enter preview text"),
)

# The field whose presence means the page finished mounting. Same role as an activity's
# ready_selector in rs_scrape: without it a hydration lag reads as an empty book.
READY_PLACEHOLDER = FIELDS[0][1]


def book_url(book_id: str) -> str:
    return f"{rs.BASE_URL}/books/{book_id}/edit"


# --------------------------------------------------------------------------------------
# Sheet layout.
# --------------------------------------------------------------------------------------


def ensure_columns(ws: Worksheet, add_missing: bool, dry_run: bool) -> dict[str, int]:
    """Guarantee every FIELDS column exists, returning a fresh header map.

    New columns are inserted straight after `id` rather than appended, so the Err
    block stays at the right-hand edge where the QA skill's `prepare` expects to
    extend it. Every lookup in both scripts is by header name, so shifting the
    existing columns right is safe.

    A dry run inserts them in memory too -- the rest of the run needs somewhere to
    report against -- but says so, since nothing reaches disk.
    """
    headers = rs.header_map(ws)
    missing = [name for name, _ in FIELDS if name.casefold() not in headers]
    if not missing:
        return headers

    if not add_missing:
        raise rs.SheetLayoutError(
            f"sheet {ws.title!r} has no {', '.join(missing)} column(s). "
            "Re-run without --no-add-columns to have them inserted."
        )

    at = rs.column_for(ws, headers, rs.ID_HEADER) + 1
    ws.insert_cols(at, amount=len(missing))
    for offset, name in enumerate(missing):
        ws.cell(row=1, column=at + offset, value=name)
    verb = "would add" if dry_run else "added"
    print(f"  + {ws.title}: {verb} column(s) {', '.join(missing)} after {rs.ID_HEADER!r}")
    return rs.header_map(ws)


# --------------------------------------------------------------------------------------
# Scrape.
# --------------------------------------------------------------------------------------


def read_fields(page: Page) -> dict[str, str]:
    """The value of each configured field, keyed by column name.

    Read with input_value(): these are controlled React fields, so inner_text() and
    the value attribute both return the server-rendered blank, not what is on screen.
    A field the book never filled in is legitimately empty and reported as such.
    """
    values: dict[str, str] = {}
    for column, placeholder in FIELDS:
        box = page.get_by_placeholder(placeholder, exact=True).first
        values[column] = rs._clean(box.input_value()) if box.count() else ""
    return values


def scrape_book(
    page: Page,
    ws: Worksheet,
    headers: dict[str, int],
    book_id: str,
    creds: tuple[str, str],
    dry_run: bool,
) -> int:
    """Fill this book's empty field cells. Returns how many cells were written."""
    row = rs.find_row(ws, headers, book_id)
    if row is None:
        print(f"  - id {book_id}: not in sheet {SHEET!r}, skipping")
        return 0

    columns = {column: rs.column_for(ws, headers, column) for column, _ in FIELDS}
    targets = {
        column: index
        for column, index in columns.items()
        if not rs._clean_cell(ws.cell(row=row, column=index).value)
    }
    if not targets:
        print(f"  = id {book_id}: both fields already recorded, skipping")
        return 0

    url = book_url(book_id)
    rs.polite_goto(page, url)

    # A session can expire part-way through a long run; without this every remaining
    # book lands on the login page and is silently reported as empty.
    if "/login" in page.url:
        print("  (session expired, logging in again)")
        rs.login(page, *creds)
        rs.polite_goto(page, url)

    try:
        page.get_by_placeholder(READY_PLACEHOLDER, exact=True).first.wait_for(
            timeout=rs.SELECTOR_TIMEOUT_MS
        )
    except PlaywrightTimeout:
        print(f"  - id {book_id}: book edit form never appeared, skipping")
        rs._screenshot(page, f"BOOK-{book_id}")
        return 0

    values = read_fields(page)
    if not any(values.values()):
        print(f"  - id {book_id}: both fields empty on the page, nothing to write")
        rs._screenshot(page, f"BOOK-{book_id}")
        return 0

    written = []
    for column, target_column in targets.items():
        value = values[column]
        if not value:
            print(f"  . id {book_id}: {column} is empty on the page, leaving blank")
            continue
        if dry_run:
            print(f"  . id {book_id}: {column} = {_preview(value)}")
        else:
            ws.cell(row=row, column=target_column, value=value)
        written.append(column)

    if written and not dry_run:
        print(f"  + id {book_id}: wrote {', '.join(written)}")
    return len(written)


def _preview(text: str, width: int = 60) -> str:
    return repr(text if len(text) <= width else text[:width] + "...")


# --------------------------------------------------------------------------------------
# Run flow. Mirrors rs_scrape.main so a long run behaves identically.
# --------------------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--workbook",
        type=Path,
        default=rs.WORKBOOK_PATH,
        help="workbook to write into (default: RS_WORKBOOK from .env)",
    )
    parser.add_argument(
        "--books",
        default="",
        help="comma-separated book ids (default: every id in the OEC sheet)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="report every intended write, save nothing",
    )
    parser.add_argument(
        "--no-add-columns",
        action="store_true",
        help="fail if Main_Character / Prev_Text are missing instead of adding them",
    )
    parser.add_argument(
        "--headed", action="store_true", help="watch the browser work"
    )
    return parser.parse_args()


def main() -> None:
    sys.stdout.reconfigure(line_buffering=True)
    args = parse_args()

    # rs_scrape's workbook helpers read this module-level path, so point them here.
    rs.WORKBOOK_PATH = args.workbook

    creds = rs.credentials()
    wb = rs.open_target_workbook()
    if SHEET not in wb.sheetnames:
        sys.exit(
            f"workbook has no {SHEET!r} sheet (sheets found: {', '.join(wb.sheetnames)})"
        )
    ws = wb[SHEET]

    if args.dry_run:
        print("DRY RUN: no cells will be written and the workbook will not be saved.\n")

    try:
        headers = ensure_columns(
            ws, add_missing=not args.no_add_columns, dry_run=args.dry_run
        )
    except rs.SheetLayoutError as exc:
        sys.exit(f"\n{exc}")

    book_ids = [b.strip() for b in args.books.split(",") if b.strip()] or (
        rs.sheet_book_ids(ws, headers)
    )

    print(
        f"Pacing: {rs.REQUEST_DELAY_S:.1f}s between page loads, "
        f"stopping after {rs.MAX_CONSECUTIVE_FAILURES} consecutive failures."
    )
    print(f"\n=== book fields -> {SHEET}: {len(book_ids)} book(s) ===")

    books_written = books_skipped = 0
    consecutive_failures = 0
    aborted: rs.RunAborted | None = None

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=not args.headed)
        state = str(rs.STORAGE_STATE) if rs.STORAGE_STATE.exists() else None
        context = browser.new_context(storage_state=state)
        rs.block_heavy_resources(context)
        page = context.new_page()
        page.set_default_navigation_timeout(rs.NAV_TIMEOUT_MS)

        try:
            rs.ensure_logged_in(page, *creds)
            rs.save_session(context)

            for book_id in book_ids:
                try:
                    written = scrape_book(
                        page, ws, headers, book_id, creds, args.dry_run
                    )
                except (PlaywrightTimeout, PlaywrightError) as exc:
                    consecutive_failures += 1
                    print(f"  ! id {book_id}: page error, skipping ({exc})")
                    books_skipped += 1
                    if consecutive_failures >= rs.MAX_CONSECUTIVE_FAILURES:
                        aborted = rs.RunAborted(
                            f"{consecutive_failures} page loads failed in a row -- "
                            "stopping so the server isn't hammered further. "
                            "Progress so far has been saved."
                        )
                        break
                    continue

                consecutive_failures = 0
                if written:
                    books_written += 1
                    if not args.dry_run:
                        rs.save_workbook(wb)
                else:
                    books_skipped += 1
        finally:
            context.close()
            browser.close()

    # The column insert is itself a change worth keeping even if no book yielded data.
    if not args.dry_run:
        rs.save_workbook(wb)

    print(
        f"\nDone. {books_written} book(s) updated, "
        f"{books_skipped} with nothing new to write."
    )
    if args.dry_run:
        print("(dry run -- workbook untouched)")
    if aborted:
        sys.exit(f"\nRun aborted: {aborted}")


if __name__ == "__main__":
    main()

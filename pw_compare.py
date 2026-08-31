"""Compare each book's Picture-Word Accuracy Check against its Picture-Word Match.

Both activities pick their words from the same global word list, so the two sets are
meant to agree. Where Picture-Word Match carries a word that Picture-Word Accuracy
Check does not, the student meets a word in one activity that the other never checks --
that is what this records, one row per extra word.

    uv run python pw_compare.py --scrape          # fills the cache, resumable
    uv run python pw_compare.py --write           # writes the sheet from the cache

Scraping and writing are separate steps because the scrape is long: 3s per page load,
two pages a book. The cache means an interrupted run resumes instead of starting over,
and the sheet can be rebuilt without touching the server again.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright

import rs_scrape as rs

SHEET = "PWAC - PWM"
CACHE = Path("data/pw_words.json")
WORKBOOKS = [Path("data/2plus_check.xlsx"), Path("data/2plus_check_batch2.xlsx")]
PAGES = {"PWAC": "picture-word-accuracy-check", "PWM": "picture-word-match"}
WORD_CELL, POS_CELL, DEF_CELL = 1, 2, 3
MISSING_MARKER = "Activity not found"

HEADERS = ["id", "Book", "Extra Word", "Word ID", "Part of Speech", "Definition",
           "Words in PWAC", "Words in PWM", "Note"]
WIDTHS = [9, 30, 24, 9, 14, 52, 13, 12, 34]
HDR_FILL = PatternFill("solid", start_color="FFD9E1F2")
TOP = Alignment(wrap_text=True, vertical="top")


def key(word: str) -> str:
    """Comparison key for a word tile: case and inflection formatting vary between the
    two activities even when they are the same word."""
    head = re.split(r"[,(]", word, maxsplit=1)[0]
    return re.sub(r"[^a-z0-9 ]+", "", head.lower()).strip()


def extract_words(page) -> list[dict[str, str]]:
    """The selected words, from the one table both edit pages render.

    With the search box empty the table holds only the words already chosen for the
    activity, which is exactly the set being compared.
    """
    tables = page.locator("table")
    if not tables.count():
        return []
    items = []
    for row in tables.first.locator("tbody tr").all():
        cells = row.locator("td")
        if cells.count() <= DEF_CELL:
            continue
        word = rs._clean(cells.nth(WORD_CELL).inner_text())
        if not word:
            continue
        items.append({
            "word": word,
            "pos": rs._clean(cells.nth(POS_CELL).inner_text()),
            "definition": rs._clean(cells.nth(DEF_CELL).inner_text()),
        })
    return items


def load_cache() -> dict:
    if CACHE.exists():
        return json.loads(CACHE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict) -> None:
    CACHE.parent.mkdir(parents=True, exist_ok=True)
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=1), encoding="utf-8")


def book_ids() -> list[str]:
    """Every book id in either workbook, in workbook order, without duplicates."""
    from openpyxl import load_workbook

    seen, out = set(), []
    for path in WORKBOOKS:
        ws = load_workbook(path, read_only=True, data_only=True)["Vocab"]
        header = {str(c.value).strip(): i for i, c in
                  enumerate(next(ws.iter_rows(max_row=1)))}
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[header["id"]]
            if v is None:
                continue
            bid = str(int(v)) if isinstance(v, float) else str(v).strip()
            if bid and bid not in seen:
                seen.add(bid)
                out.append(bid)
    return out


def scrape(limit: int | None) -> None:
    sys.stdout.reconfigure(line_buffering=True)
    cache = load_cache()
    ids = [b for b in book_ids() if b not in cache]
    if limit:
        ids = ids[:limit]
    print(f"{len(cache)} book(s) already cached, {len(ids)} to fetch "
          f"({2 * len(ids)} page loads, about "
          f"{2 * len(ids) * rs.REQUEST_DELAY_S / 60:.0f} min at current pacing)")

    consecutive = 0
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=rs.HEADLESS)
        state = str(rs.STORAGE_STATE) if rs.STORAGE_STATE.exists() else None
        context = browser.new_context(storage_state=state)
        rs.block_heavy_resources(context)
        page = context.new_page()
        page.set_default_navigation_timeout(rs.NAV_TIMEOUT_MS)
        rs.ensure_logged_in(page, *rs.credentials())
        rs.save_session(context)

        try:
            for n, bid in enumerate(ids, 1):
                entry: dict = {}
                failed = False
                for label, segment in PAGES.items():
                    url = f"{rs.BASE_URL}/activities/{bid}/{segment}/edit"
                    try:
                        rs.polite_goto(page, url)
                    except (PlaywrightTimeout, PlaywrightError) as exc:
                        print(f"  ! {bid} {label}: {type(exc).__name__}")
                        failed = True
                        break
                    # The table mounts after the activity loads; a missing activity
                    # renders an error instead, so wait for either rather than a
                    # fixed sleep.
                    try:
                        page.wait_for_selector(
                            f"table, :text('{MISSING_MARKER}')",
                            timeout=rs.SELECTOR_TIMEOUT_MS)
                    except PlaywrightTimeout:
                        pass
                    body = page.inner_text("body")
                    if MISSING_MARKER in body:
                        entry[label] = None          # activity does not exist
                    else:
                        entry[label] = extract_words(page)
                if failed:
                    consecutive += 1
                    if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                        print(f"\n{consecutive} failures in a row -- stopping. "
                              f"{len(cache)} book(s) cached.")
                        break
                    continue
                consecutive = 0
                cache[bid] = entry
                shape = " ".join(
                    f"{k}={'-' if v is None else len(v)}" for k, v in entry.items())
                print(f"  [{n}/{len(ids)}] {bid}: {shape}")
                if n % 10 == 0:
                    save_cache(cache)
        finally:
            save_cache(cache)
            browser.close()
    print(f"\ncached {len(cache)} book(s) -> {CACHE}")


def ensure_sheet(wb):
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment = Font(bold=True), HDR_FILL, TOP
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c - 1]
    ws.freeze_panes = "A2"
    return ws


def write() -> None:
    from openpyxl import load_workbook

    import manual_qa
    import word_ids

    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    # The extra words are global words too, so give them the same id the Vocab sheet
    # carries -- without it there is no way to tell which sense of the word is meant.
    tiers = (word_ids.build_index(word_ids.load_global())
             if word_ids.CACHE.exists() else None)

    def word_id(item: dict) -> str:
        if not tiers:
            return ""
        wid, _, _ = word_ids.resolve(item["word"], item["definition"], tiers,
                                     item["pos"])
        return wid or ""
    cache = load_cache()
    if not cache:
        sys.exit(f"{CACHE} is empty -- run with --scrape first.")
    _, rows = manual_qa.load_books()
    titles = {b: t for b, t, _ in rows}

    totals = {"extra": 0, "both": 0, "pwac only": 0, "pwm only": 0, "neither": 0}
    for path in WORKBOOKS:
        wb = load_workbook(path)
        ids = [b for b in book_ids()
               if rs.find_row(wb["Vocab"], rs.header_map(wb["Vocab"]), b) is not None]
        ws = ensure_sheet(wb)
        written = 0
        for bid in ids:
            entry = cache.get(bid)
            if entry is None:
                continue
            pwac, pwm = entry.get("PWAC"), entry.get("PWM")
            if pwac is None and pwm is None:
                totals["neither"] += 1
                continue
            if pwm is None:
                totals["pwac only"] += 1
                row = ws.max_row + 1
                for c, v in enumerate([bid, titles.get(bid, ""), "", "", "", "",
                                       len(pwac), "-",
                                       "This book has a Picture-Word Accuracy Check "
                                       "but no Picture-Word Match."], 1):
                    ws.cell(row, c, v).alignment = TOP
                written += 1
                continue
            if pwac is None:
                totals["pwm only"] += 1
                for item in pwm:
                    row = ws.max_row + 1
                    for c, v in enumerate([bid, titles.get(bid, ""), item["word"],
                                           word_id(item), item["pos"],
                                           item["definition"], "-", len(pwm),
                                           "This book has no Picture-Word Accuracy "
                                           "Check at all, so every Picture-Word Match "
                                           "word is unchecked."], 1):
                        ws.cell(row, c, v).alignment = TOP
                    written += 1
                    totals["extra"] += 1
                continue
            totals["both"] += 1
            have = {key(i["word"]) for i in pwac}
            for item in pwm:
                if key(item["word"]) in have:
                    continue
                row = ws.max_row + 1
                for c, v in enumerate([bid, titles.get(bid, ""), item["word"],
                                       word_id(item), item["pos"],
                                       item["definition"], len(pwac), len(pwm),
                                       "In Picture-Word Match but not in Picture-Word "
                                       "Accuracy Check."], 1):
                    ws.cell(row, c, v).alignment = TOP
                written += 1
                totals["extra"] += 1
        wb.save(path)
        print(f"{path.name}: {written} row(s) on '{SHEET}'")
    print("\nbooks with both activities:", totals["both"],
          "| PWAC only:", totals["pwac only"],
          "| PWM only:", totals["pwm only"],
          "| neither:", totals["neither"])
    print("extra words recorded:", totals["extra"])


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--scrape", action="store_true")
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--limit", type=int, help="only fetch this many books")
    args = ap.parse_args()
    if args.scrape:
        scrape(args.limit)
    if args.write:
        write()
    if not (args.scrape or args.write):
        ap.error("choose --scrape, --write, or both")


if __name__ == "__main__":
    main()

"""Colour every row of a report that has changed since a baseline copy.

Written for the case where someone has reviewed a report by hand and then the underlying
data moved on: they need to see what is new without re-reading what they already signed
off.

Rows are matched on (book id, details, suggested fix) rather than on the field name,
because the Vocab sheet was renumbered when it switched to the book word list -- a
finding that was W3 may now be W7 while being the same finding about the same word.
Matching on the field would paint hundreds of untouched rows as changed and bury the
handful that genuinely are.

    green   a finding that did not exist at the baseline
    yellow  the same finding, but its field was renumbered
    grey    a finding that was in the baseline and is now gone (listed, not painted)

    uv run python highlight_changes.py --baseline <dir> --reports references \\
        --prefix 0820_error-reports
"""

from __future__ import annotations

import argparse
import collections
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

NEW = PatternFill("solid", start_color="FFD9EAD3")        # green
RENUMBERED = PatternFill("solid", start_color="FFFFF2CC")  # yellow
MERGED = PatternFill("solid", start_color="FFE4D7F5")      # lavender
LEGEND = [("New since your review", "FFD9EAD3"),
          ("Same finding, field renumbered", "FFFFF2CC"),
          ("Merged in from the second batch", "FFE4D7F5")]
VIEWS = ("recurring", "by-type")


def norm(text) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def detail_rows(path: Path):
    """(key -> field) for every detail row, plus the row positions for painting."""
    wb = load_workbook(path)
    for name in wb.sheetnames[1:]:
        ws = wb[name]
        idx = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)
               if ws.cell(4, c).value}
        if "Book ID" not in idx:
            continue
        for r in range(5, ws.max_row + 1):
            cell = lambda h: norm(ws.cell(r, idx[h]).value) if h in idx else ""
            if not cell("Book ID"):
                continue
            yield ws, r, idx, (cell("Book ID"), cell("Details"),
                               cell("Suggested Fix")), cell("Field")


def issue_of(feedback: str) -> str:
    m = re.search(r"^Issue: (.*)$", feedback or "", re.M)
    return norm(m.group(1)) if m else ""


def long_rows(path: Path):
    wb = load_workbook(path)
    ws = wb["Content"]
    for r in range(6, ws.max_row + 1):
        fb = ws.cell(r, 8).value
        if fb:
            yield ws, r, (norm(ws.cell(r, 5).value), issue_of(fb))


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--baseline", type=Path, required=True)
    ap.add_argument("--reports", type=Path, required=True)
    ap.add_argument("--prefix", required=True)
    ap.add_argument("--flag-books", default="",
                    help="comma-separated book ids to paint as merged-in")
    args = ap.parse_args()
    flagged = {b.strip() for b in args.flag_books.split(",") if b.strip()}

    # --- the two grouped views ---------------------------------------------------
    base_field: dict[tuple, str] = {}
    for view in VIEWS:
        p = args.baseline / f"{args.prefix}_{view}.xlsx"
        if p.exists():
            for _, _, _, key, field in detail_rows(p):
                base_field.setdefault(key, field)

    counts: collections.Counter = collections.Counter()
    seen_now: set = set()
    # The long view identifies a book by title only, so the id has to come from the
    # grouped views, which carry both. Two books sharing a title would resolve to
    # whichever id was seen first; the run prints its merged-in count so a mismatch
    # against the number of flagged findings is visible rather than silent.
    long_book_ids: dict[str, str] = {}
    for view in VIEWS:
        path = args.reports / f"{args.prefix}_{view}.xlsx"
        if not path.exists():
            continue
        wb = load_workbook(path)
        painted = 0
        for name in wb.sheetnames[1:]:
            ws = wb[name]
            idx = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1)
                   if ws.cell(4, c).value}
            if "Book ID" not in idx:
                continue
            for r in range(5, ws.max_row + 1):
                cell = lambda h: norm(ws.cell(r, idx[h]).value) if h in idx else ""
                if not cell("Book ID"):
                    continue
                key = (cell("Book ID"), cell("Details"), cell("Suggested Fix"))
                seen_now.add(key)
                long_book_ids.setdefault(cell("Book"), cell("Book ID"))
                if cell("Book ID") in flagged:
                    fill, kind = MERGED, "merged"
                elif key not in base_field:
                    fill, kind = NEW, "new"
                elif base_field[key] != cell("Field"):
                    fill, kind = RENUMBERED, "renumbered"
                else:
                    continue
                counts[(view, kind)] += 1
                painted += 1
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill
        # Legend above the summary sheet's own header block.
        summary = wb[wb.sheetnames[0]]
        for i, (label, colour) in enumerate(LEGEND):
            cell = summary.cell(1 + i, 11, label)
            cell.fill = PatternFill("solid", start_color=colour)
        wb.save(path)
        print(f"{path.name}: painted {painted} row(s)")

    # --- the long view -----------------------------------------------------------
    base_long = set()
    p = args.baseline / f"{args.prefix}_long.xlsx"
    if p.exists():
        base_long = {key for _, _, key in long_rows(p)}
    path = args.reports / f"{args.prefix}_long.xlsx"
    if path.exists():
        wb = load_workbook(path)
        ws = wb["Content"]
        painted = 0
        for r in range(6, ws.max_row + 1):
            fb = ws.cell(r, 8).value
            if not fb:
                continue
            book_id = long_book_ids.get(norm(ws.cell(r, 5).value))
            merged = book_id in flagged if book_id else False
            if not merged and (norm(ws.cell(r, 5).value), issue_of(fb)) in base_long:
                continue
            painted += 1
            for c in range(1, 15):
                ws.cell(r, c).fill = MERGED if merged else NEW
        for i, (label, colour) in enumerate(LEGEND):
            cell = ws.cell(1 + i, 11, label)
            cell.fill = PatternFill("solid", start_color=colour)
        wb.save(path)
        print(f"{path.name}: painted {painted} row(s)")

    gone = set(base_field) - seen_now
    tally = lambda kind: sum(v for (_, k), v in counts.items() if k == kind)
    print(f"\nnew: {tally('new')} | renumbered: {tally('renumbered')} "
          f"| merged in from another batch: {tally('merged')}")
    print(f"in the baseline but not in the current report: {len(gone)}")
    for key in sorted(gone)[:15]:
        print(f"   book {key[0]}: {key[1][:70]}")


if __name__ == "__main__":
    main()

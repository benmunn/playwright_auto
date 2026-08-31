"""Give every vocabulary entry the id of the global word it came from.

The same spelling is linked to more than one book while meaning different things --
`bark` the sound and `bark` on a tree -- and only the global word id tells those apart.
The workbook records the word and its definition, so the id has to be looked up.

Rather than searching the list once per definition, this walks the whole global list
once (28k words, 100 to a page) and matches locally. That is a tenth of the requests, it
can be re-matched without touching the server again, and it makes the near-misses
visible: a book whose definition has drifted from the global one shows up as a fallback
match rather than silently resolving to the wrong sense.

    uv run python word_ids.py --scrape          # fill data/global_words.json, resumable
    uv run python word_ids.py --match           # report how the two sides line up
    uv run python word_ids.py --match --write   # add the id columns

Matching runs in tiers, strictest first, and every entry records which tier resolved it
so a reviewer can audit the weak ones.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright

import rs_scrape as rs

CACHE = Path("data/global_words.json")
PAGE_SIZE = 100
ID_CELL, WORD_CELL, POS_CELL, DEF_CELL = 0, 2, 3, 4
# The list also carries both translations, which is why they are read here rather
# than from the per-word edit form: 284 pages against nearly four thousand.
KOR_CELL, VIE_CELL = 5, 6
# An empty translation renders as a dash.
EMPTY = {"", "-", "–", "—"}


def norm(text: str) -> str:
    """Whitespace- and unicode-normalised, but still case-sensitive."""
    text = unicodedata.normalize("NFKC", text or "")
    return re.sub(r"\s+", " ", text.replace("’", "'").replace("“", '"')
                  .replace("”", '"')).strip()


def loose(text: str) -> str:
    """As above, then case-folded and stripped of punctuation, for the fallback tiers."""
    return re.sub(r"[^a-z0-9 ]+", "", norm(text).lower()).strip()


# A handful of global entries abbreviate the part of speech; the workbooks never do.
POS_ALIAS = {"n": "noun", "v": "verb", "adj": "adjective", "adv": "adverb",
             "pre": "preposition", "phr": "phrase", "phrasal verb": "verb"}


def pos_key(text: str) -> str:
    p = loose(text)
    return POS_ALIAS.get(p, p)


# --------------------------------------------------------------------------------------
# Scrape
# --------------------------------------------------------------------------------------


# Reading a cell at a time costs one round trip to the browser each, which at six cells
# across a hundred rows is six hundred per page and dwarfs the polite delay between
# pages. The whole table comes back in one call instead.
ROWS_JS = """() => Array.from(document.querySelectorAll('table tbody tr'),
    tr => Array.from(tr.querySelectorAll('td'), td => td.innerText))"""


def read_page(page) -> list[dict[str, str]]:
    out = []
    for cells in page.evaluate(ROWS_JS):
        if len(cells) <= DEF_CELL:
            continue
        wid = rs._clean(cells[ID_CELL])
        if not wid.isdigit():
            continue

        def cell(i, cells=cells):
            if len(cells) <= i:
                return ""
            text = rs._clean(cells[i])
            return "" if text in EMPTY else text

        out.append({"id": wid,
                    "word": cell(WORD_CELL),
                    "pos": cell(POS_CELL),
                    "definition": cell(DEF_CELL),
                    "kor": cell(KOR_CELL),
                    "vie": cell(VIE_CELL)})
    return out


def total_pages(page) -> int:
    body = page.inner_text("body")
    m = re.search(r"of\s+([\d,]+)", body)
    if not m:
        sys.exit("could not read the total from the page -- has the layout changed?")
    total = int(m.group(1).replace(",", ""))
    print(f"{total:,} word(s) in the global list")
    return (total + PAGE_SIZE - 1) // PAGE_SIZE


def scrape(limit: int | None) -> None:
    sys.stdout.reconfigure(line_buffering=True)
    cache = json.loads(CACHE.read_text(encoding="utf-8")) if CACHE.exists() else {}
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=rs.HEADLESS)
        state = str(rs.STORAGE_STATE) if rs.STORAGE_STATE.exists() else None
        context = browser.new_context(storage_state=state)
        rs.block_heavy_resources(context)
        page = context.new_page()
        page.set_default_navigation_timeout(rs.NAV_TIMEOUT_MS)
        rs.ensure_logged_in(page, *rs.credentials())
        rs.save_session(context)

        rs.polite_goto(page, f"{rs.BASE_URL}/words?page=1&pageSize={PAGE_SIZE}")
        page.wait_for_selector("table tbody tr", timeout=rs.SELECTOR_TIMEOUT_MS)
        pages = total_pages(page)
        todo = [n for n in range(1, pages + 1)
                if str(n) not in cache
                or any("kor" not in r for r in cache[str(n)])]
        if limit:
            todo = todo[:limit]
        print(f"{len(cache)} page(s) cached, {len(todo)} to fetch "
              f"(about {len(todo) * rs.REQUEST_DELAY_S / 60:.0f} min)")

        consecutive = 0
        try:
            for i, n in enumerate(todo, 1):
                url = f"{rs.BASE_URL}/words?page={n}&pageSize={PAGE_SIZE}"
                try:
                    rs.polite_goto(page, url)
                    page.wait_for_selector("table tbody tr",
                                           timeout=rs.SELECTOR_TIMEOUT_MS)
                except (PlaywrightTimeout, PlaywrightError) as exc:
                    consecutive += 1
                    print(f"  ! page {n}: {type(exc).__name__}")
                    if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                        print("stopping: too many failures in a row")
                        break
                    continue
                consecutive = 0
                rows = read_page(page)
                if not rows:
                    print(f"  ! page {n}: no rows read")
                    continue
                cache[str(n)] = rows
                if i % 10 == 0 or i == len(todo):
                    CACHE.write_text(json.dumps(cache, ensure_ascii=False),
                                     encoding="utf-8")
                    print(f"  [{i}/{len(todo)}] page {n}: {len(rows)} row(s), "
                          f"{sum(len(v) for v in cache.values()):,} cached")
        finally:
            CACHE.parent.mkdir(parents=True, exist_ok=True)
            CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
            browser.close()
    print(f"\n{sum(len(v) for v in cache.values()):,} word(s) cached in {CACHE}")


# --------------------------------------------------------------------------------------
# Match
# --------------------------------------------------------------------------------------


def load_global() -> list[dict[str, str]]:
    if not CACHE.exists():
        sys.exit(f"{CACHE} is missing -- run with --scrape first.")
    cache = json.loads(CACHE.read_text(encoding="utf-8"))
    seen, out = set(), []
    for rows in cache.values():
        for row in rows:
            if row["id"] not in seen:
                seen.add(row["id"])
                out.append(row)
    return out


def build_index(words: list[dict[str, str]]) -> dict[str, dict]:
    """One lookup per tier. A key that is ambiguous is dropped from its tier rather than
    resolved arbitrarily -- an arbitrary id here is exactly the confusion being fixed."""
    tiers: dict[str, dict[tuple, list]] = {
        "exact": {}, "loose": {}, "prefix": {}, "word": {}}
    for w in words:
        tiers["exact"].setdefault((norm(w["word"]), norm(w["definition"])), []).append(w)
        tiers["loose"].setdefault((loose(w["word"]), loose(w["definition"])), []).append(w)
        tiers["prefix"].setdefault(loose(w["word"]), []).append(w)
        tiers["word"].setdefault(loose(w["word"]), []).append(w)
    return tiers


def resolve(word: str, definition: str, tiers: dict,
            pos: str = "") -> tuple[str | None, str, int]:
    """(id, how it was matched, how many candidates the tier held).

    Where a tier holds more than one candidate the part of speech breaks the tie: the
    global list carries 522 word-and-definition pairs twice, and all but 15 of those are
    one entry filed under two different parts of speech.
    """

    def settle(cands: list, how: str,
               tie: bool = True) -> tuple[str | None, str, int] | None:
        if len(cands) == 1:
            return cands[0]["id"], how, 1
        if len(cands) > 1:
            same = [w for w in cands if pos and pos_key(w["pos"]) == pos_key(pos)]
            if tie and len(same) == 1:
                return same[0]["id"], f"{how}, part of speech breaking the tie", 1
            return None, f"ambiguous on {how}", len(cands)
        return None

    got = settle(tiers["exact"].get((norm(word), norm(definition)), []),
                 "word + definition")
    if got:
        return got

    got = settle(tiers["loose"].get((loose(word), loose(definition)), []),
                 "word + definition, ignoring case and punctuation")
    if got:
        return got

    # The book's definition sometimes carries an extra clause the global one lacks.
    cands = tiers["prefix"].get(loose(word), [])
    d = loose(definition)
    got = settle([w for w in cands if d and (d.startswith(loose(w["definition"]))
                                             or loose(w["definition"]).startswith(d))],
                 "word, and one definition is the start of the other")
    if got:
        return got

    # No tie-break here: with the definition already failed to match, word and part of
    # speech alone would pick a sense rather than confirm one.
    got = settle(cands, "word only, unique in the global list", tie=False)
    if got:
        return got
    return None, "no global word with this spelling", 0


# --------------------------------------------------------------------------------------
# Writing the id back
# --------------------------------------------------------------------------------------

WORKBOOKS = [Path("data/2plus_check.xlsx"), Path("data/2plus_check_batch2.xlsx")]


def vocab_entries(ws, headers):
    """(row, slot, word, part of speech, definition) for every filled vocabulary slot."""
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, headers["id"]).value
        if v is None:
            continue
        for s in range(1, rs.MAX_SLOTS + 1):
            if f"w{s}" not in headers:
                break
            word = str(ws.cell(r, headers[f"w{s}"]).value or "").strip()
            if not word:
                continue
            yield (r, s, word,
                   str(ws.cell(r, headers[f"pos{s}"]).value or "").strip(),
                   str(ws.cell(r, headers[f"def{s}"]).value or "").strip())


def match(write: bool) -> None:
    from openpyxl import load_workbook

    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    tiers = build_index(load_global())
    print(f"{sum(len(v) for v in tiers['exact'].values()):,} global word(s) indexed")

    how: dict[str, int] = {}
    unresolved: list[tuple] = []
    resolved: dict[tuple[str, str], str] = {}
    for path in WORKBOOKS:
        if not path.exists():
            continue
        wb = load_workbook(path)
        ws = wb["Vocab"]
        headers = rs.header_map(ws)
        # One id column per word slot, inserted once, right after the last Err set so the
        # sheet's existing column positions do not move.
        if write:
            for s in range(1, rs.MAX_SLOTS + 1):
                if f"w{s}" not in headers:
                    break
                name = f"WID{s}"
                if name.casefold() not in headers:
                    col = ws.max_column + 1
                    ws.cell(1, col, name)
                    headers[name.casefold()] = col
        for r, s, word, pos, definition in vocab_entries(ws, headers):
            wid, reason, n = resolve(word, definition, tiers, pos)
            how[reason] = how.get(reason, 0) + 1
            if wid:
                resolved[(word, definition)] = wid
                if write:
                    ws.cell(r, headers[f"wid{s}"], int(wid))
            else:
                unresolved.append((path.name, r, s, word, definition, reason, n))
        if write:
            wb.save(path)
            print(f"saved {path}")

    total = sum(how.values())
    print(f"\n{total:,} vocabulary entr(ies)")
    for reason, n in sorted(how.items(), key=lambda kv: -kv[1]):
        print(f"  {n:>5}  {reason}")
    print(f"\n{len(unresolved)} entr(ies) with no id")
    for row in unresolved[:25]:
        print(f"  {row[0]} row{row[1]} W{row[2]}: {row[3]!r}")
        print(f"      {row[5]} ({row[6]} candidate(s)) :: {row[4][:70]}")
    Path("data/word_id_unresolved.json").write_text(
        json.dumps([{"workbook": w, "row": r, "slot": s, "word": word,
                     "definition": d, "why": why, "candidates": n}
                    for w, r, s, word, d, why, n in unresolved],
                   ensure_ascii=False, indent=1), encoding="utf-8")
    if not write:
        print("\nDRY RUN -- no columns written. Add --write.")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--scrape", action="store_true")
    ap.add_argument("--match", action="store_true")
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--limit", type=int, help="only fetch this many pages")
    args = ap.parse_args()
    if args.scrape:
        scrape(args.limit)
    if args.match:
        match(args.write)
    if not (args.scrape or args.match):
        ap.error("choose --scrape, --match, or both")


if __name__ == "__main__":
    main()

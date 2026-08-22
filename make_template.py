"""Generate an empty workbook with the sheets and headers rs_scrape.py expects.

    uv run make_template.py                          -> data/activities_template.xlsx
    uv run make_template.py --ids 697 --out x.xlsx    -> same, with 697 in the id column

Headers only, no data. rs_scrape.py locates columns by name, so the column order here is
purely for human readability -- W1/POS1/DEF1 are grouped together rather than blocked as
W1..W10, POS1..POS10 for the same reason.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import book_fields
import rs_scrape

# Column families per sheet, in the order they should appear. Each entry is a tuple of
# prefixes that repeat together for every slot 1..SLOTS.
SHEET_LAYOUTS: dict[str, tuple[tuple[str, ...], ...]] = {
    "OEC": (("Q",),),
    "CC": (("Q",), ("A",)),
    "Vocab": (("W", "POS", "DEF", "SENT"),),
    "TMC": (("Q", "AnsA", "AnsB", "AnsC", "AnsD", "Correct"),),
}

# Single-value columns, written between `id` and the first slot family. Unlike the
# families above these do not repeat per slot: one book has one preview text. Taken
# from book_fields so the layout and the scraper cannot drift apart.
FLAT_COLUMNS: dict[str, tuple[str, ...]] = {
    "OEC": tuple(name for name, _ in book_fields.FIELDS),
}

# Slots to generate per sheet. rs_scrape.MAX_SLOTS is only an upper bound for reading;
# these are the widths actually written, so a sheet is no wider than its content needs.
DEFAULT_SLOTS = 10
SHEET_SLOTS = {"Vocab": 25}   # longest word list seen is 21; leave headroom
SLOTS = DEFAULT_SLOTS
ID_HEADER = rs_scrape.ID_HEADER

# Rough widths per prefix, so the file is usable without manual resizing.
WIDTHS = {
    ID_HEADER: 10,
    "Main_Character": 28,
    "Prev_Text": 60,
    "Q": 52,
    "A": 20,
    "W": 18,
    "POS": 8,
    "DEF": 46,
    "SENT": 52,
    "AnsA": 26,
    "AnsB": 26,
    "AnsC": 26,
    "AnsD": 26,
    "Correct": 9,
}

HEADER_FILL = PatternFill("solid", start_color="FFD9E1F2")
HEADER_FONT = Font(bold=True)


def headers_for(sheet: str) -> list[str]:
    """The full header row: id, this sheet's flat columns, then the slot families."""
    row = [ID_HEADER, *FLAT_COLUMNS.get(sheet, ())]
    slots = SHEET_SLOTS.get(sheet, DEFAULT_SLOTS)
    for family in SHEET_LAYOUTS[sheet]:
        for slot in range(1, slots + 1):
            row.extend(f"{prefix}{slot}" for prefix in family)
    return row


def build(ids: list[str]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in SHEET_LAYOUTS:
        ws = wb.create_sheet(sheet)
        row = headers_for(sheet)
        ws.append(row)

        for column, header in enumerate(row, start=1):
            cell = ws.cell(row=1, column=column)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            prefix = header.rstrip("0123456789") or header
            ws.column_dimensions[get_column_letter(column)].width = WIDTHS.get(prefix, 18)

        for book_id in ids:
            ws.append([book_id])

        # Keep the header row, the id column and any flat columns visible while
        # scrolling -- on OEC the preview text is context for every question to its
        # right, so it has to stay on screen.
        frozen = 1 + len(FLAT_COLUMNS.get(sheet, ()))
        ws.freeze_panes = f"{get_column_letter(frozen + 1)}2"

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("data/activities_template.xlsx"),
        help="output path (default: data/activities_template.xlsx)",
    )
    parser.add_argument(
        "--ids",
        nargs="*",
        default=[],
        metavar="ID",
        help="book ids to seed into the id column of every sheet",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="overwrite the output file if it already exists",
    )
    args = parser.parse_args()

    if args.out.exists() and not args.force:
        raise SystemExit(
            f"{args.out} already exists. Pass --force to overwrite, or choose --out."
        )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb = build([str(book_id) for book_id in args.ids])
    wb.save(args.out)

    sheets = ", ".join(f"{name} ({len(headers_for(name))} cols)" for name in SHEET_LAYOUTS)
    print(f"Wrote {args.out}")
    print(f"  sheets: {sheets}")
    print(f"  slots per family: 1..{SLOTS}")
    if args.ids:
        print(f"  seeded ids: {', '.join(str(i) for i in args.ids)}")


if __name__ == "__main__":
    main()

"""Render the three report views: one row per error, grouped by kind, grouped by type.

All three are built from the same list of findings, so they must always agree; the
verification step for any report run is that the three views hold identical row multisets
and that the workbook they came from agrees with them.

Layout follows the RnD x TFT feedback spreadsheet: colour-code legend, header on row 5,
and the same column names, so a reader who knows that file knows these.
"""
from __future__ import annotations

import collections
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as L

from . import vocab_changes
from .blurbs import BLURBS
from .classify import BY_CODE, RULES

Q = lambda s: f'"{s}"' if s else "(blank)"
HDR_FILL = PatternFill("solid", start_color="FFD9E1F2")
TOP = Alignment(wrap_text=True, vertical="top")
ORDER = {"OEC": 0, "CC": 1, "Vocab": 2, "TMC": 3, "LR": 4, "LRA": 5}
LEGEND = [(1, 5, "Urgent; manageable by R&D", "FF00FFFF", False),
          (1, 8, "Less urgent; manageable by R&D", "FFD9EAD3", False),
          (1, 9, "Mentioned to Devs", "FFE6B8AF", True),
          (2, 5, "Urgent; requires dev action", "FFF6B26B", False),
          (2, 8, "Less urgent; requires dev action", "FFFCE5CD", False),
          (2, 9, "Resolved", "FFCFE2F3", True)]

# Per-activity detail columns, then the tail every sheet shares.
COLS = {
    "Vocab": [("Word", lambda f: f["ctx"].get("word", "")),
              ("Word ID", lambda f: f["ctx"].get("word_id", "")),
              ("Part of Speech", lambda f: f["ctx"].get("pos", "")),
              ("Definition", lambda f: f["ctx"].get("definition", "")),
              ("Story Sentence", lambda f: f["ctx"].get("sentence", ""))],
    "CC": [("Sentence", lambda f: f["ctx"].get("sentence", "")),
           ("Bank Answer", lambda f: f["ctx"].get("answer", ""))],
    "TMC": [("Question", lambda f: f["ctx"].get("question", "")),
            ("Answer Key", lambda f: f["ctx"].get("key", "")),
            ("Options A-D", lambda f: "\n".join(
                f"{p}. {f['ctx'].get('options', {}).get(p, '')}" for p in "ABCD"))],
    "OEC": [("Preview Text", lambda f: f["ctx"].get("prev_text", "")),
            ("Main Character(s)", lambda f: f["ctx"].get("main_character", ""))],
    # Listen & Read has no scraped content to hang a finding off, so the page and the
    # sentence the reviewer quoted stand in for it.
    "LR": [("Page", lambda f: f["ctx"].get("page", "")),
           ("Sentence", lambda f: f["ctx"].get("text", ""))],
    "LRA": [("Page", lambda f: f["ctx"].get("page", "")),
            ("Sentence", lambda f: f["ctx"].get("text", ""))],
}
# Book first on every detail sheet: the reader works a book at a time, so the column
# they scan should not be buried past the activity content.
HEAD = [("Book", lambda f: f["book"]), ("Book ID", lambda f: f["book_id"])]
TAIL = [("Field", lambda f: f["target"]), ("Current Text", lambda f: f["current"]),
        ("Suggested Fix", lambda f: f["fix"]), ("Details", lambda f: f["details"]),
        ("Reason", lambda f: f["type"])]
WIDTH = {"Page": 8, "Text": 52, "Word": 20, "Word ID": 9, "Part of Speech": 13, "Definition": 46, "Story Sentence": 50,
         "Sentence": 52, "Bank Answer": 16, "Question": 46, "Answer Key": 10,
         "Options A-D": 42, "Preview Text": 46, "Main Character(s)": 24, "Field": 11,
         "Current Text": 44, "Suggested Fix": 44, "Details": 46, "Book ID": 9,
         "Book": 30, "Reason": 17}

SUM_HDR = ["Date", "ID", "Name", "Level", "Book", "Error Type", "Relevant Activity",
           "Feedback", "Relevant Books", "Relevant Images/Videos", "Response Name",
           "Response", "Occurrences", "Books Affected", "Detail Sheet"]
SUM_W = [8.1, 6.0, 13.2, 11.2, 20.0, 24.0, 27.2, 78.0, 40.0, 20.0, 18.1, 20.0,
         12.0, 14.0, 30.0]
LONG_HDR = SUM_HDR[:12] + ["1열", "Device, Browser"]
LONG_W = [8.1, 6.0, 13.2, 11.2, 29.9, 11.1, 27.2, 90.0, 28.0, 47.2, 18.1, 38.6, 19.0, 17.2]

# What each activity is called on a tab. The workbook sheet is "OEC", but the team
# calls the activity OEQ, so that is what the report is labelled with.
TAB_NAME = {"OEC": "OEQ"}

# Short tab names for the "Other" sub-splits; the full label still heads each sheet.
OTHER_SHORT = {"W05": "repeats the word", "W15": "factual error",
               "T15": "factual-logic fault", "T01": "stem has no blank",
               "T14": "misspelling", "T12": "duplicate options", "T13": "wrong content",
               "C04": "duplicate sentence", "C10": "other content fault",
               "O09": "other fault"}


def _legend(ws, with_phonics: bool = False) -> None:
    ws.cell(1, 4, "Color Code:").font = Font(bold=True)
    rows = LEGEND + ([(3, 8, "Phonics issues; manageable/requires action by ReadingStar",
                       "FFEAD1DC", False)] if with_phonics else [])
    for r, c, text, rgb, bold in rows:
        cell = ws.cell(r, c, text)
        cell.fill = PatternFill("solid", start_color=rgb)
        cell.font = Font(bold=bold)


def _detail_sheet(wb, tab: str, heading: str, items: list[dict]) -> None:
    det = wb.create_sheet(tab)
    det.cell(1, 1, heading).font = Font(bold=True, size=12)
    det.cell(2, 1, f"{len(items)} occurrence(s) across "
                   f"{len({f['book_id'] for f in items})} book(s) - "
                   f"{items[0]['activity']} [{items[0]['phase']}]")
    cols = HEAD + COLS[items[0]["sheet"]] + TAIL
    for c, (h, _) in enumerate(cols, 1):
        cell = det.cell(4, c, h)
        cell.font, cell.fill, cell.alignment = Font(bold=True), HDR_FILL, TOP
        det.column_dimensions[L(c)].width = WIDTH.get(h, 18)
    for i, f in enumerate(sorted(items, key=lambda f: (f["book"].lower(),
                                                       f["n"] or 0, f["target"]))):
        for c, (h, get) in enumerate(cols, 1):
            det.cell(5 + i, c, get(f)).alignment = TOP
    det.freeze_panes = "A5"


def _example(items: list[dict]) -> tuple[str, str]:
    """A short instance to illustrate the group, when none is written by hand.

    Shortest wins: two long sentences that differ by one word make the reader hunt for
    the difference, which is the opposite of what an example is for.
    """
    usable = [f for f in items if f["fix"] and len(f["current"]) > 12]
    pick = min(usable or items, key=lambda f: len(f["current"]))
    return pick["current"], pick["fix"]


def group_feedback(key, title: str, items: list[dict]) -> str:
    """The Feedback cell for one by-type group: what the error is, then one example."""
    if key in BLURBS:
        desc, bullets, current, suggested = BLURBS[key]
    else:
        kinds = sorted({BY_CODE[f["kind"]][4] for f in items})
        desc, bullets = kinds[0], kinds[1:]
        current, suggested = _example(items)
    lines = [f"[{items[0]['phase']}]", f"{items[0]['activity']} - {title}",
             "Link to sheet: ", "", desc]
    lines += [f"- {b}" for b in bullets]
    lines += ["", "Example:", f"- current: {current}", f"- suggested: {suggested}"]
    return "\n".join(lines)


def _summary(wb, title: str, entries) -> None:
    """entries: (label, extra Feedback line, tab name, items, group key or None).

    A group key means the row gets the by-type Feedback layout; None keeps the older
    one, which the recurring view still uses.
    """
    ws = wb.active
    ws.title = title
    _legend(ws)
    for c, h in enumerate(SUM_HDR, 1):
        cell = ws.cell(5, c, h)
        cell.font, cell.fill, cell.alignment = Font(bold=True), HDR_FILL, TOP
        ws.column_dimensions[L(c)].width = SUM_W[c - 1]
    ws.freeze_panes = "A6"
    for i, (label, extra, tab, items, key) in enumerate(entries):
        r = 6 + i
        ex = max(items, key=lambda f: len(f["fix"]))
        books = sorted({(f["book_id"], f["book"]) for f in items},
                       key=lambda b: b[1].lower())
        ws.cell(r, 6, label).alignment = TOP
        ws.cell(r, 7, items[0]["activity"]).alignment = TOP
        body = (group_feedback(key, label, items) if key is not None else "\n".join([
            f"[{items[0]['phase']}]", items[0]["activity"], "", label, "",
            f"Recorded {len(items)} time(s) across {len(books)} book(s).", extra, "",
            f"Example - {ex['book']} ({ex['target']})",
            f"Current: {Q(ex['current'])}", f"Issue: {ex['details']}",
            f"Suggested fix: {Q(ex['fix'])}", "",
            f"Every instance is listed on sheet: {tab}"]))
        ws.cell(r, 8, body).alignment = TOP
        ws.cell(r, 9, "\n".join(t for _, t in books)).alignment = TOP
        ws.cell(r, 13, len(items)).alignment = TOP
        ws.cell(r, 14, len(books)).alignment = TOP
        ws.cell(r, 15, tab).alignment = TOP
        ws.row_dimensions[r].height = 150


def feedback(f: dict) -> str:
    """The TFT-style Feedback cell: where the error is, what it is, what to write."""
    ctx, n, tgt = f["ctx"], f["n"], f["target"]
    lines = [f"[{f['phase']}]", f["activity"], ""]
    if f["sheet"] == "TMC":
        if tgt.startswith("AnsAll"):
            lines += [f"Q{n}: {Q(ctx.get('question'))}",
                      f"All four options ({tgt}) - {f['type']}"]
        elif f["prefix"].startswith("Ans"):
            lines += [f"Q{n}: {Q(ctx.get('question'))}",
                      f"Option {f['prefix'][-1]} ({tgt}) - {f['type']}"]
        elif f["prefix"] == "Correct":
            lines += [f"Q{n}: {Q(ctx.get('question'))}",
                      f"Answer key ({tgt}, currently {ctx.get('key') or '-'}) - {f['type']}"]
        else:
            lines += [f"Question {n} ({tgt}) - {f['type']}"]
    elif f["sheet"] == "CC":
        if f["prefix"] == "A":
            lines += [f"Word bank item {n} ({tgt}) - {f['type']}",
                      f"Used in sentence {n}: {Q(ctx.get('sentence'))}"]
        else:
            lines += [f"Sentence {n} ({tgt}) - {f['type']}",
                      f"Bank answer: {Q(ctx.get('answer'))}"]
    elif f["sheet"] == "Vocab":
        # The global word id, so the fixer edits the entry this book is actually
        # linked to rather than another book's word of the same spelling.
        wid = f" [word #{ctx['word_id']}]" if ctx.get("word_id") else ""
        lines += [f"Entry {n}: {Q(ctx.get('word'))} ({ctx.get('pos') or '-'}){wid}",
                  f"Definition: {Q(ctx.get('definition'))}"]
        if ctx.get("sentence"):
            lines.append(f"Story sentence: {Q(ctx['sentence'])}")
        lines.append(f"Target {tgt} - {f['type']}")
    elif f["sheet"] in ("LR", "LRA"):
        page = ctx.get("page")
        lines += [f"Page {page}" if page else "Page not recorded",
                  f"{f['type']}"]
    else:
        label = {"Main_Character": "Main character list",
                 "Prev_Text": "Preview text"}.get(tgt, f"Question {n} ({tgt})")
        lines += [f"{label} - {f['type']}"]
        if ctx.get("prev_text"):
            lines.append(f"Preview text: {Q(ctx['prev_text'])}")
        if ctx.get("main_character"):
            lines.append(f"Main character(s): {ctx['main_character']}")
    lines += ["", f"Current: {Q(f['current'])}", f"Issue: {f['details']}",
              f"Suggested fix: {Q(f['fix'])}"]
    return "\n".join(lines)


def render_long(data: list[dict], path) -> pathlib.Path:
    """One row per error, sorted by book so a fixer can work a book at a time."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Content"
    _legend(ws, with_phonics=True)
    for c, h in enumerate(LONG_HDR, 1):
        cell = ws.cell(5, c, h)
        cell.font, cell.alignment = Font(bold=True), TOP
        ws.column_dimensions[L(c)].width = LONG_W[c - 1]
    rows = sorted(data, key=lambda f: (f["book"].lower(), ORDER[f["sheet"]],
                                       f["n"] or 0, f["target"]))
    for i, f in enumerate(rows):
        r = 6 + i
        ws.cell(r, 5, f["book"]).alignment = TOP
        ws.cell(r, 7, f["activity"]).alignment = TOP
        ws.cell(r, 8, feedback(f)).alignment = TOP
    ws.freeze_panes = "A6"
    wb.save(path)
    return pathlib.Path(path)


def render_recurring(data: list[dict], path) -> pathlib.Path:
    """One row per recurring kind of error, with every instance on its own sheet."""
    by_kind = collections.defaultdict(list)
    for f in data:
        by_kind[f["kind"]].append(f)
    wb = Workbook()
    entries = []
    for code, sh, ty, pat, label, tab in RULES:
        items = by_kind.get(code)
        if items:
            reasons = sorted({f["type"] for f in items})
            entries.append((label, f"Logged under: {', '.join(reasons)}", tab,
                            items, None))
    _summary(wb, "Recurring Errors", entries)
    for label, extra, tab, items, _ in entries:
        _detail_sheet(wb, tab, label, items)
    wb.save(path)
    return pathlib.Path(path)


VC_COLS = [("Book", 34), ("Book ID", 11), ("word_id", 10), ("current_word", 20),
           ("fixed_word", 20), ("current_pos", 13), ("fixed_pos", 13),
           ("current_def", 52), ("fixed_def", 52), ("current_kor", 34),
           ("fixed_kor", 34), ("current_vie", 38), ("fixed_vie", 38),
           ("Error Types", 22), ("Details", 74)]


def _vocab_sheet(wb, name: str, at: int, heading: str, note: str, rows: list[dict],
                 titles: dict) -> None:
    ws = wb.create_sheet(name, at)
    ws.cell(1, 1, heading).font = Font(bold=True, size=12)
    ws.cell(2, 1, note)
    for c, (h, w) in enumerate(VC_COLS, 1):
        cell = ws.cell(4, c, h)
        cell.font, cell.fill, cell.alignment = Font(bold=True), HDR_FILL, TOP
        ws.column_dimensions[L(c)].width = w
    for i, r in enumerate(rows):
        values = [r["book_names"], r["book_ids"], r["id"], r["word"], r["fixed_word"],
                  r["pos"], r["fixed_pos"], r["definition"], r["fixed_def"],
                  r["kor"], r["fixed_kor"], r["vie"], r["fixed_vie"],
                  ", ".join(r["types"]), vocab_changes.details(r, titles)]
        for c, v in enumerate(values, 1):
            ws.cell(5 + i, c, v).alignment = TOP
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{L(len(VC_COLS))}{4 + len(rows)}"


def render_vocab_changes(wb, data: list[dict], uses: dict, titles: dict,
                         glossary: dict | None = None) -> int:
    """Every change wanted to the global vocabulary entries, on two sheets.

    Entries that one edit can serve sit on the first sheet, one row each. Entries whose
    books mean different things by the same word cannot be edited in one go, so they get
    their own sheet: each row there is one sense, and acting on it means creating a
    separate word entry rather than editing the shared one.
    """
    rows = vocab_changes.build_rows(data, uses, glossary)
    for r in rows:
        r["book_names"] = "; ".join(titles.get(b, b) for b in r["books"])
        r["book_ids"] = ", ".join(r["books"])
    single = [r for r in rows if not r["split"]]
    split = [r for r in rows if r["split"]]
    _vocab_sheet(wb, "Vocab Changes", 1, "Vocabulary changes, one row per word entry",
                 f"{len(single)} word entries, one edit each. A blank fixed_* cell means "
                 f"that field needs no change. Entries whose books disagree about the "
                 f"meaning are not here -- see the 'Vocab Changes - Split' sheet.",
                 single, titles)
    _vocab_sheet(wb, "Vocab Changes - Split", 2,
                 "Vocabulary entries used in more than one sense",
                 f"{len(split)} rows covering {len({r['word_id'] for r in split})} word "
                 f"entries. Each id is suffixed A, B, C ... one row per sense. These "
                 f"cannot be fixed by editing the shared entry: split it into separate "
                 f"entries and relink each book to the right one. Rows with no suggested "
                 f"fix are senses nobody flagged that the other row's fix would break.",
                 split, titles)
    return len(rows)


def render_by_type(data: list[dict], path, uses: dict | None = None,
                   titles: dict | None = None,
                   glossary: dict | None = None) -> pathlib.Path:
    """One row per logged type. Only "Other" is split further, since it is a catch-all."""
    groups = collections.defaultdict(list)
    for f in data:
        key = ((f["sheet"], "Other", f["kind"]) if f["type"] == "Other"
               else (f["sheet"], f["type"], None))
        groups[key].append(f)
    wb = Workbook()
    entries, seen = [], {}
    for k in sorted(groups, key=lambda k: (ORDER[k[0]], k[1] == "Other", k[1], k[2] or "")):
        sheet_key, ty, code = k
        items = groups[k]
        title = ty if code is None else f"Other - {BY_CODE[code][4]}"
        name = TAB_NAME.get(sheet_key, sheet_key)
        tab = (f"{name} {ty}" if code is None
               else f"{name} Other - {OTHER_SHORT.get(code, code)}")[:31]
        seen[tab] = seen.get(tab, 0) + 1
        if seen[tab] > 1:
            tab = f"{tab[:28]}~{seen[tab]}"
        kinds = sorted({BY_CODE[f["kind"]][4] for f in items})
        entries.append((title, "Covers: " + "; ".join(kinds), tab, items, k))
    _summary(wb, "Errors by Type", entries)
    if uses is not None:
        render_vocab_changes(wb, data, uses, titles or {}, glossary)
    for title, extra, tab, items, _ in entries:
        _detail_sheet(wb, tab, f"{items[0]['sheet']} - {title}", items)
    wb.save(path)
    return pathlib.Path(path)

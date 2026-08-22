"""Scrape authored activity content from the Reading Space admin site into a workbook.

Logs into admin.reading-space.com, walks the activity-edit page for each book id, and
records the questions / answers / vocabulary it finds into the matching sheet of an
existing .xlsx file. See the constants block below for everything you'd want to change.

The workbook is never created or restructured: sheets and header rows must already
exist. Columns are located by header name, and only empty cells are written.
"""

from __future__ import annotations

import os
import sys
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Locator, Page, TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright

# --------------------------------------------------------------------------------------
# Section A -- configuration. This is the part you edit.
#
# Read at import time so .env can supply values, and so overriding any of these
# from another module (as the tests do) still works.
# --------------------------------------------------------------------------------------

load_dotenv()

BASE_URL = "https://admin.reading-space.com"

# Path to the existing workbook holding the OEC / CC / Vocab / TMC sheets. Set
# RS_WORKBOOK in .env to point at your own file; the default is only a fallback.
WORKBOOK_PATH = Path(os.getenv("RS_WORKBOOK") or "data/activities.xlsx")

# Which activities to run, in order. Any subset of ACTIVITIES' keys.
ACTIVITIES_TO_RUN: tuple[str, ...] = ("OEC", "CC", "Vocab", "TMC")

# None = every id in that sheet's own "id" column. Otherwise a list, e.g. ["101", "102"].
BOOK_IDS: list[str] | None = None

# Upper bound on numbered slots per column family (Q1..Q10, W1..W20, ...). This is only
# a cap: slot_count() stops at the first slot the sheet has no header for, so raising it
# lets a wider sheet through without widening the narrow ones. Vocab needs the headroom
# because a book's word list runs to about twenty entries, against the eight the Word
# Accuracy Check activity used to expose.
MAX_SLOTS = 30

# False = watch the browser work. Useful while confirming selectors against the live site.
HEADLESS = True

# True = report every intended write but save nothing. Use this for the first run.
DRY_RUN = False

# True = a value already recorded for that book is not written again, so re-running the
# script is a no-op instead of appending a second copy into the next free slots. Set to
# False for literal append-at-first-empty-slot behaviour.
SKIP_ALREADY_RECORDED = True

# Cached login session, so re-runs skip the login form. Delete it to force a fresh login.
STORAGE_STATE = Path(".auth/state.json")

# Screenshots of pages that failed to yield data.
DEBUG_DIR = Path("debug")

# Every OEC prediction textarea carries this placeholder.
OEC_PLACEHOLDER = "What do you think this story will be about?"

# The book word list at /books/<id>/words. Cell order is:
# 0 select | 1 image | 2 word | 3 part of speech | 4 definition | 5 ko | 6 vi |
# 7 sentences | 8 actions. The sentence input is matched by placeholder, as a prefix --
# the live text ends in an ellipsis that is not worth depending on.
VOCAB_WORD_CELL, VOCAB_POS_CELL, VOCAB_DEF_CELL = 2, 3, 4
VOCAB_SENTENCE_PLACEHOLDER = "Enter a sentence"

NAV_TIMEOUT_MS = 30_000
SELECTOR_TIMEOUT_MS = 15_000

# The admin panel is a React app. Immediately after domcontentloaded the markup
# exists but React has not attached state to it yet: form values read as blank and
# the login button stays disabled. Every navigation therefore waits for the network
# to go quiet before anything is read or typed.
HYDRATION_TIMEOUT_MS = 15_000

ID_HEADER = "id"


# --------------------------------------------------------------------------------------
# Politeness. The admin panel is a low-resource back-office tool, so this script is
# deliberately slow: one page at a time, a minimum gap between page loads, a long pause
# before retrying a failure, and a hard stop if the server looks like it is struggling.
#
# Raise REQUEST_DELAY_S if the server still feels stressed. The whole run is sequential,
# so throughput is roughly (number of books x number of activities x REQUEST_DELAY_S).
# --------------------------------------------------------------------------------------

# Minimum seconds between the start of one page load and the start of the next.
REQUEST_DELAY_S = 3.0

# Extra seconds to wait after a page load fails, before the single retry.
RETRY_BACKOFF_S = 20.0

# How many times to retry one failed page load. 1 keeps load low; 0 disables retries.
MAX_RETRIES = 1

# Give up on the whole run after this many books fail back-to-back, rather than keep
# hammering a server that is already in trouble.
MAX_CONSECUTIVE_FAILURES = 3

# Resource types not fetched at all. A page load is many requests, not one, and none of
# these affect the text this script reads -- blocking them cuts server hits substantially.
# Stylesheets and scripts are deliberately NOT blocked: the admin panel is a React app,
# so its scripts must run for the form values to exist.
BLOCKED_RESOURCES = frozenset({"image", "media", "font"})


# --------------------------------------------------------------------------------------
# Section B -- selectors.
#
# Structural locators are used wherever the element is identifiable that way. The two
# class-string selectors below are the exceptions: the CC answer pills and the TMC
# question cards carry no role, label or stable id, so their Tailwind classes are the
# only handle. They are written as class-subset selectors, so an added utility class
# won't break the match -- but a restyle will, and these are the first place to look
# if an activity suddenly extracts nothing.
# --------------------------------------------------------------------------------------

CC_PILL = "div.flex.items-center.gap-2.rounded-full.bg-neutral-100"
TMC_CARD = "div.flex.flex-col.gap-4.rounded-lg.bg-white.p-6.shadow-md"
TMC_QUESTION_LABEL = "label.flex.flex-col.gap-2"
TMC_ANSWER_LIST = "div.flex.flex-col.gap-2 div.flex.flex-col.gap-3"

# Answer options in page order; the column suffix is the prefix minus "Ans".
TMC_ANSWER_PREFIXES = ("AnsA", "AnsB", "AnsC", "AnsD")

# The <select> value that marks an option as the answer key. Its siblings are
# "closestAnswer" and "offTopic", which we deliberately do not record.
TMC_CORRECT_VALUE = "correct"


# --------------------------------------------------------------------------------------
# Section C -- extractors.
#
# Each returns one dict per item, keyed by column prefix, in the order the items appear
# on the page. Controlled React inputs are read with input_value(): text_content() on a
# <textarea> returns the server-rendered default, which is typically empty.
# --------------------------------------------------------------------------------------


def _clean(value: str | None) -> str:
    return (value or "").strip()


def extract_oec_questions(page: Page) -> list[dict[str, str]]:
    """Every prediction textarea on the open-ended-questions page."""
    items = []
    for box in page.get_by_placeholder(OEC_PLACEHOLDER).all():
        text = _clean(box.input_value())
        if text:
            items.append({"Q": text})
    return items


def extract_cc_questions(page: Page) -> list[dict[str, str]]:
    """Every question textarea on the context-clue page."""
    items = []
    for box in page.locator("textarea").all():
        text = _clean(box.input_value())
        if text:
            items.append({"Q": text})
    return items


def extract_cc_answers(page: Page) -> list[dict[str, str]]:
    """The answer pills on the context-clue page -- read-only, so inner_text()."""
    items = []
    for pill in page.locator(CC_PILL).all():
        label = pill.locator("span").first
        if label.count() == 0:
            continue
        text = _clean(label.inner_text())
        if text:
            items.append({"A": text})
    return items


def extract_vocab(page: Page) -> list[dict[str, str]]:
    """Word / part of speech / definition / story sentence from the book's word list.

    Read from /books/<id>/words, not from the Word Accuracy Check activity page. The
    activity page shows only the subset of words that activity uses -- 8 of book 798's
    20 -- and carries neither the story sentence nor, for many books, anything at all.
    The word list is the authoritative record, so it is what gets scraped.

    Columns are: select | image | word | part of speech | definition | Korean |
    Vietnamese | sentences | actions. The sentence is a controlled React input, so it
    is read with input_value() and located by placeholder rather than by cell index.
    """
    table = page.locator("table").first
    items = []
    for row in table.locator("tbody tr").all():
        cells = row.locator("td")
        if cells.count() <= VOCAB_DEF_CELL:
            continue
        word = _clean(cells.nth(VOCAB_WORD_CELL).inner_text())
        if not word:
            continue
        box = row.get_by_placeholder(VOCAB_SENTENCE_PLACEHOLDER).first
        items.append(
            {
                "W": word,
                "POS": _clean(cells.nth(VOCAB_POS_CELL).inner_text()),
                "DEF": _clean(cells.nth(VOCAB_DEF_CELL).inner_text()),
                "SENT": _clean(box.input_value()) if box.count() else "",
            }
        )
    return items


def extract_tmc_questions(page: Page) -> list[dict[str, str]]:
    """One question plus its four answer options per question card."""
    items = []
    for card in page.locator(TMC_CARD).all():
        question_box = card.locator(f"{TMC_QUESTION_LABEL} textarea").first
        if question_box.count() == 0:
            continue
        question = _clean(question_box.input_value())
        if not question:
            continue

        options, correct = _tmc_answer_values(card)
        item = {"Q": question, "Correct": correct}
        for prefix, value in zip(TMC_ANSWER_PREFIXES, options):
            item[prefix] = value
        items.append(item)
    return items


def _tmc_answer_values(card: Locator) -> tuple[list[str], str]:
    """The card's four answer texts, plus the letter of the one marked correct.

    Each option carries a <select> whose options are correct / closestAnswer / offTopic.
    Which one is chosen appears nowhere in the markup -- there is no `selected`
    attribute -- so it has to be read as a DOM property via input_value(), the same rule
    that applies to the textareas. That costs no extra page load.

    The letter is only reported when exactly one option is marked. "correct" is the
    first <option> in the list, so a dropdown the app never set reads as "correct" by
    default -- and an untouched card therefore shows every option as correct. Treating
    that as an answer key would invent one, so anything other than exactly one match
    yields "" and says so.
    """
    answer_list = card.locator(TMC_ANSWER_LIST).first
    values: list[str] = []
    marked: list[str] = []
    if answer_list.count():
        for index, option in enumerate(answer_list.locator("> div").all()):
            field = option.locator("input").first
            if field.count() == 0:
                continue
            values.append(_clean(field.input_value()))

            marker = option.locator("select").first
            if (
                index < len(TMC_ANSWER_PREFIXES)
                and marker.count()
                and _clean(marker.input_value()) == TMC_CORRECT_VALUE
            ):
                marked.append(TMC_ANSWER_PREFIXES[index].removeprefix("Ans"))

    if len(marked) > 1:
        print(
            f"    (ambiguous answer key: {', '.join(marked)} all marked correct -- "
            "recording none)"
        )
    correct = marked[0] if len(marked) == 1 else ""
    return (values + ["", "", "", ""])[:4], correct


# --------------------------------------------------------------------------------------
# Section B (cont.) -- activity registry.
#
# A field group is one family of columns that must stay row-aligned. `anchor` is the
# prefix whose first empty slot decides where the group starts writing; `prefixes` are
# all the columns the group fills at that slot.
#
# CC has two groups because its questions and its answer pills are independent
# sequences of possibly different length, each seeking its own first empty slot. Vocab and
# TMC each have one group, because W3/POS3/DEF3 come from a single table row and must
# land at the same slot number.
# --------------------------------------------------------------------------------------


@dataclass(frozen=True)
class FieldGroup:
    anchor: str
    prefixes: tuple[str, ...]
    extract: Callable[[Page], list[dict[str, str]]]
    # Locator that must appear before extraction; its absence means "no data here".
    ready_selector: str


@dataclass(frozen=True)
class Activity:
    key: str
    sheet: str
    url_segment: str
    groups: tuple[FieldGroup, ...]
    # Most activities live under /activities/<book>/<segment>/edit. Vocab does not: its
    # content is the book's own word list, at a book-scoped URL with no segment.
    url_pattern: str = "{base}/activities/{book}/{segment}/edit"

    def url(self, book_id: str) -> str:
        return self.url_pattern.format(
            base=BASE_URL, book=book_id, segment=self.url_segment
        )


ACTIVITIES: dict[str, Activity] = {
    "OEC": Activity(
        key="OEC",
        sheet="OEC",
        url_segment="open-ended-questions",
        groups=(
            FieldGroup(
                anchor="Q",
                prefixes=("Q",),
                extract=extract_oec_questions,
                ready_selector="textarea",
            ),
        ),
    ),
    "CC": Activity(
        key="CC",
        sheet="CC",
        url_segment="context-clue",
        groups=(
            FieldGroup(
                anchor="Q",
                prefixes=("Q",),
                extract=extract_cc_questions,
                ready_selector="textarea",
            ),
            FieldGroup(
                anchor="A",
                prefixes=("A",),
                extract=extract_cc_answers,
                ready_selector=CC_PILL,
            ),
        ),
    ),
    # Vocab is the book's own word list, not an activity page. The Word Accuracy Check
    # and Word Meaning Match activities each show a subset of it; the list itself is the
    # record, and the only place the story sentence exists.
    "Vocab": Activity(
        key="Vocab",
        sheet="Vocab",
        url_segment="",
        url_pattern="{base}/books/{book}/words",
        groups=(
            FieldGroup(
                anchor="W",
                prefixes=("W", "POS", "DEF", "SENT"),
                extract=extract_vocab,
                ready_selector="table tbody tr",
            ),
        ),
    ),
    "TMC": Activity(
        key="TMC",
        sheet="TMC",
        url_segment="text-multiple-choice",
        groups=(
            FieldGroup(
                anchor="Q",
                prefixes=("Q", *TMC_ANSWER_PREFIXES, "Correct"),
                extract=extract_tmc_questions,
                ready_selector=TMC_CARD,
            ),
        ),
    ),
}


# --------------------------------------------------------------------------------------
# Section D -- workbook access.
# --------------------------------------------------------------------------------------


class SheetLayoutError(RuntimeError):
    """The workbook is missing a sheet or a column the activity needs."""


def header_map(ws: Worksheet) -> dict[str, int]:
    """Row 1 as {header text: column index}, case-insensitive and whitespace-trimmed."""
    headers: dict[str, int] = {}
    for column, cell in enumerate(ws[1], start=1):
        label = _clean(str(cell.value)) if cell.value is not None else ""
        if label:
            headers.setdefault(label.casefold(), column)
    return headers


def column_for(ws: Worksheet, headers: dict[str, int], name: str) -> int:
    column = headers.get(name.casefold())
    if column is None:
        raise SheetLayoutError(
            f"sheet {ws.title!r} has no {name!r} column "
            f"(headers found: {', '.join(sorted(headers)) or 'none'})"
        )
    return column


def sheet_book_ids(ws: Worksheet, headers: dict[str, int]) -> list[str]:
    """Every non-empty value in the sheet's id column, in row order."""
    id_column = column_for(ws, headers, ID_HEADER)
    ids = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=id_column).value
        book_id = _cell_key(value)
        if book_id:
            ids.append(book_id)
    return ids


def find_row(ws: Worksheet, headers: dict[str, int], book_id: str) -> int | None:
    """The row whose id cell matches book_id, comparing as trimmed strings."""
    id_column = column_for(ws, headers, ID_HEADER)
    wanted = _cell_key(book_id)
    for row in range(2, ws.max_row + 1):
        if _cell_key(ws.cell(row=row, column=id_column).value) == wanted:
            return row
    return None


def _cell_key(value: object) -> str:
    """Normalise an id cell for comparison, so 101, 101.0 and " 101 " all match."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def slot_count(ws: Worksheet, headers: dict[str, int], group: FieldGroup) -> int:
    """How many slots this sheet actually provides for every prefix in the group.

    A sheet laid out Q1..Q8 gets eight slots rather than a hard failure at Q9; a sheet
    missing slot 1 outright is a layout error, since there is nowhere to write.
    """
    count = 0
    for slot in range(1, MAX_SLOTS + 1):
        if all(f"{prefix}{slot}".casefold() in headers for prefix in group.prefixes):
            count = slot
        else:
            break
    if count == 0:
        missing = ", ".join(f"{prefix}1" for prefix in group.prefixes)
        raise SheetLayoutError(
            f"sheet {ws.title!r} is missing the first slot column(s): {missing}"
        )
    return count


def first_empty_slot(
    ws: Worksheet, row: int, headers: dict[str, int], anchor: str, slots: int
) -> int | None:
    """Lowest slot number 1..slots whose anchor cell is empty, else None."""
    for slot in range(1, slots + 1):
        column = column_for(ws, headers, f"{anchor}{slot}")
        if not _clean_cell(ws.cell(row=row, column=column).value):
            return slot
    return None


def _clean_cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def recorded_values(
    ws: Worksheet, row: int, headers: dict[str, int], anchor: str, slots: int
) -> list[str]:
    """Non-empty anchor values already present in this row, in slot order."""
    values = []
    for slot in range(1, slots + 1):
        column = column_for(ws, headers, f"{anchor}{slot}")
        value = _clean_cell(ws.cell(row=row, column=column).value)
        if value:
            values.append(value)
    return values


def drop_already_recorded(
    items: list[dict[str, str]], anchor: str, existing: list[str]
) -> list[dict[str, str]]:
    """Filter out items whose anchor value is already recorded for this book.

    Matched as a multiset, not a set: if the page legitimately shows the same answer
    twice and only one copy is recorded, the second copy still gets written.
    """
    unmatched = Counter(existing)
    fresh = []
    for item in items:
        key = item.get(anchor, "").strip()
        if unmatched.get(key, 0) > 0:
            unmatched[key] -= 1
            continue
        fresh.append(item)
    return fresh


def write_group(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    group: FieldGroup,
    items: list[dict[str, str]],
    label: str,
) -> tuple[int, int]:
    """Write items into consecutive slots from the anchor's first empty one.

    Returns (items written, items dropped for lack of free slots).
    """
    if not items:
        return 0, 0

    slots = slot_count(ws, headers, group)

    if SKIP_ALREADY_RECORDED:
        existing = recorded_values(ws, row, headers, group.anchor, slots)
        fresh = drop_already_recorded(items, group.anchor, existing)
        if len(fresh) < len(items):
            print(
                f"  = {label}: {len(items) - len(fresh)} of {len(items)} value(s) "
                "already recorded, skipping those"
            )
        items = fresh
        if not items:
            return 0, 0

    start = first_empty_slot(ws, row, headers, group.anchor, slots)
    if start is None:
        print(
            f"  ! {label}: {group.anchor}1-{group.anchor}{slots} are all full, "
            f"dropping {len(items)} value(s): {_preview(items, group.anchor)}"
        )
        return 0, len(items)

    free = slots - start + 1
    written, dropped = items[:free], items[free:]

    for offset, item in enumerate(written):
        slot = start + offset
        for prefix in group.prefixes:
            column = column_for(ws, headers, f"{prefix}{slot}")
            value = item.get(prefix, "")
            if DRY_RUN:
                print(f"  . {label}: {prefix}{slot} = {value!r}")
            else:
                ws.cell(row=row, column=column, value=value)

    if not DRY_RUN and written:
        span = f"{group.anchor}{start}"
        if len(written) > 1:
            span += f"-{group.anchor}{start + len(written) - 1}"
        print(f"  + {label}: wrote {len(written)} item(s) into {span}")

    if dropped:
        print(
            f"  ! {label}: only {free} slot(s) free from {group.anchor}{start}, "
            f"dropping {len(dropped)} value(s): {_preview(dropped, group.anchor)}"
        )
    return len(written), len(dropped)


def _preview(items: list[dict[str, str]], anchor: str, width: int = 40) -> str:
    parts = []
    for item in items:
        text = item.get(anchor, "")
        parts.append(repr(text if len(text) <= width else text[:width] + "..."))
    return ", ".join(parts)


# --------------------------------------------------------------------------------------
# Throttled navigation.
# --------------------------------------------------------------------------------------


class RunAborted(RuntimeError):
    """The server looks unhealthy, so the run stopped early rather than pile on."""


_last_request_at = 0.0


def block_heavy_resources(context) -> None:
    """Stop the browser fetching assets this script never reads."""
    def handler(route):
        if route.request.resource_type in BLOCKED_RESOURCES:
            route.abort()
        else:
            route.continue_()

    context.route("**/*", handler)


def throttle() -> None:
    """Sleep until at least REQUEST_DELAY_S has passed since the last page load."""
    global _last_request_at
    wait = REQUEST_DELAY_S - (time.monotonic() - _last_request_at)
    if wait > 0:
        time.sleep(wait)
    _last_request_at = time.monotonic()


def await_hydration(page: Page) -> None:
    """Wait for the SPA to finish mounting before reading or typing anything.

    Falls through on timeout rather than failing: the per-activity ready_selector wait
    is the real gate, and a page that polls in the background never goes fully idle.
    """
    try:
        page.wait_for_load_state("networkidle", timeout=HYDRATION_TIMEOUT_MS)
    except PlaywrightTimeout:
        pass


def polite_goto(page: Page, url: str) -> None:
    """Navigate with a minimum inter-request gap and one slow retry on failure.

    Every navigation in this script goes through here, so the pacing holds no matter
    which activity or code path triggered the load.
    """
    for attempt in range(MAX_RETRIES + 1):
        throttle()
        try:
            page.goto(url, wait_until="domcontentloaded")
            await_hydration(page)
            return
        except (PlaywrightTimeout, PlaywrightError) as exc:
            if attempt >= MAX_RETRIES:
                raise
            print(
                f"    page load failed ({type(exc).__name__}); "
                f"backing off {RETRY_BACKOFF_S:.0f}s before one retry"
            )
            time.sleep(RETRY_BACKOFF_S)


# --------------------------------------------------------------------------------------
# Browser session.
# --------------------------------------------------------------------------------------


def credentials() -> tuple[str, str]:
    username = os.getenv("RS_USERNAME", "").strip()
    password = os.getenv("RS_PASSWORD", "")
    if not username or not password:
        sys.exit(
            "RS_USERNAME and RS_PASSWORD are not both set.\n"
            "Copy .env.example to .env and fill in your admin credentials."
        )
    return username, password


def login(page: Page, username: str, password: str) -> None:
    print("Logging in...")
    polite_goto(page, f"{BASE_URL}/login")
    page.fill("#username", username)
    page.fill("#password", password)

    # The button is disabled until React registers both fields. Waiting on it turns a
    # hydration problem into a clear message instead of a 30s click timeout.
    try:
        page.wait_for_selector(
            "button[type=submit]:not([disabled])", timeout=SELECTOR_TIMEOUT_MS
        )
    except PlaywrightTimeout:
        sys.exit(
            "The login button never became enabled, so the credentials were never "
            "submitted. The page markup has probably changed -- check the #username / "
            "#password selectors in login()."
        )
    page.click("button[type=submit]")
    try:
        page.wait_for_url(lambda url: "/login" not in url, timeout=SELECTOR_TIMEOUT_MS)
    except PlaywrightTimeout:
        sys.exit(
            "Still on the login page after submitting -- check RS_USERNAME / "
            "RS_PASSWORD in .env."
        )
    print("Logged in.")


def ensure_logged_in(page: Page, username: str, password: str) -> None:
    """Confirm the cached session still works, logging in again if it doesn't."""
    polite_goto(page, f"{BASE_URL}/")
    if "/login" in page.url:
        login(page, username, password)


def save_session(context) -> None:
    STORAGE_STATE.parent.mkdir(parents=True, exist_ok=True)
    context.storage_state(path=str(STORAGE_STATE))


# --------------------------------------------------------------------------------------
# Run flow.
# --------------------------------------------------------------------------------------


def scrape_book(page: Page, ws: Worksheet, headers: dict[str, int],
                activity: Activity, book_id: str, creds: tuple[str, str]
                ) -> tuple[int, int]:
    """Scrape one activity page for one book. Returns (written_count, dropped_count)."""
    row = find_row(ws, headers, book_id)
    if row is None:
        print(f"  - id {book_id}: not in sheet {activity.sheet!r}, skipping")
        return 0, 0

    url = activity.url(book_id)
    polite_goto(page, url)

    # A session can expire part-way through a long run. Without this, every remaining
    # book would land on the login page and be silently reported as "no content".
    if "/login" in page.url:
        print("  (session expired, logging in again)")
        login(page, *creds)
        polite_goto(page, url)

    written = 0
    dropped = 0
    for group in activity.groups:
        label = f"id {book_id} {activity.key}/{group.anchor}"
        try:
            page.wait_for_selector(group.ready_selector, timeout=SELECTOR_TIMEOUT_MS)
        except PlaywrightTimeout:
            print(f"  - {label}: no matching content on the page, skipping")
            _screenshot(page, f"{activity.key}-{book_id}-{group.anchor}")
            continue

        items = group.extract(page)
        if not items:
            print(f"  - {label}: nothing extracted, skipping")
            _screenshot(page, f"{activity.key}-{book_id}-{group.anchor}")
            continue

        group_written, group_dropped = write_group(
            ws, row, headers, group, items, label
        )
        written += group_written
        dropped += group_dropped

    return written, dropped


def _screenshot(page: Page, name: str) -> None:
    try:
        DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(DEBUG_DIR / f"{name}.png"), full_page=True)
    except PlaywrightError as exc:
        print(f"    (could not save screenshot: {exc})")


def open_target_workbook():
    if not WORKBOOK_PATH.exists():
        sys.exit(
            f"Workbook not found: {WORKBOOK_PATH}\n"
            "Set RS_WORKBOOK in .env to the existing .xlsx holding the "
            "OEC / CC / Vocab / TMC sheets (or run make_template.py to make one)."
        )
    try:
        return load_workbook(WORKBOOK_PATH)
    except PermissionError:
        sys.exit(
            f"Cannot open {WORKBOOK_PATH} -- it is probably open in Excel. "
            "Close it and re-run."
        )


def save_workbook(wb) -> None:
    """Persist progress. Called after each book so a crash can't lose finished work."""
    try:
        wb.save(WORKBOOK_PATH)
    except PermissionError:
        sys.exit(
            f"\nCannot save {WORKBOOK_PATH} -- it is probably open in Excel. "
            "Close it and re-run; already-recorded books will be skipped."
        )


def main() -> None:
    # Progress must stay visible when output is piped to a log or a pager; without this
    # a long run looks hung for minutes while Python buffers its stdout.
    sys.stdout.reconfigure(line_buffering=True)

    creds = credentials()
    wb = open_target_workbook()

    if DRY_RUN:
        print("DRY RUN: no cells will be written and the workbook will not be saved.\n")

    books_written = books_skipped = values_dropped = 0
    consecutive_failures = 0
    aborted: RunAborted | None = None

    print(
        f"Pacing: {REQUEST_DELAY_S:.1f}s between page loads, "
        f"stopping after {MAX_CONSECUTIVE_FAILURES} consecutive failures."
    )

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        state = str(STORAGE_STATE) if STORAGE_STATE.exists() else None
        context = browser.new_context(storage_state=state)
        block_heavy_resources(context)
        page = context.new_page()
        page.set_default_navigation_timeout(NAV_TIMEOUT_MS)

        try:
            ensure_logged_in(page, *creds)
            save_session(context)

            for key in ACTIVITIES_TO_RUN:
                activity = ACTIVITIES[key]
                if activity.sheet not in wb.sheetnames:
                    raise SheetLayoutError(
                        f"workbook has no {activity.sheet!r} sheet "
                        f"(sheets found: {', '.join(wb.sheetnames)})"
                    )
                ws = wb[activity.sheet]
                headers = header_map(ws)
                book_ids = BOOK_IDS or sheet_book_ids(ws, headers)

                print(f"\n=== {activity.key}: {len(book_ids)} book(s) ===")
                for book_id in book_ids:
                    try:
                        written, dropped = scrape_book(
                            page, ws, headers, activity, book_id, creds
                        )
                    except (PlaywrightTimeout, PlaywrightError) as exc:
                        consecutive_failures += 1
                        print(f"  ! id {book_id}: page error, skipping ({exc})")
                        books_skipped += 1
                        if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                            raise RunAborted(
                                f"{consecutive_failures} page loads failed in a row -- "
                                "stopping so the server isn't hammered further. "
                                "Progress so far has been saved."
                            ) from exc
                        continue

                    consecutive_failures = 0
                    values_dropped += dropped
                    if written:
                        books_written += 1
                        if not DRY_RUN:
                            save_workbook(wb)
                    else:
                        books_skipped += 1
        except RunAborted as exc:
            aborted = exc
        finally:
            context.close()
            browser.close()

    print(
        f"\nDone. {books_written} book(s) updated, {books_skipped} with nothing new to "
        f"write, {values_dropped} value(s) dropped for lack of free slots."
    )
    if DRY_RUN:
        print("(dry run -- workbook untouched)")
    if aborted:
        sys.exit(f"\nRun aborted: {aborted}")


if __name__ == "__main__":
    main()

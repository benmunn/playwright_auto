"""Apply the vocabulary fixes from the error report to the global word list.

    uv run python word_edit.py                     # dry run: says what it would change
    uv run python word_edit.py --limit 5           # dry run over the first five
    uv run python word_edit.py --apply --limit 5   # actually save those five
    uv run python word_edit.py --apply             # the whole sheet

This is the first script here that writes. Everything else reads, so a mistake costs a
re-run; a mistake here overwrites the live word list. The guards below all exist for
that reason, and none of them should be removed to make a run go faster.

    Dry run is the default.        --apply is required before anything is saved.
    The page is checked first.     The word, part of speech and definition on the form
                                   must match what the report recorded before any field
                                   is touched. If the entry has been edited since the
                                   report was generated, the row is skipped, because the
                                   suggested fix was written against text that is gone.
    The save button proves it.     The form only enables Save once something has really
                                   changed, so an enabled button is the confirmation
                                   that the edits registered, and a disabled one means
                                   the run would have saved nothing.
    Every word is journalled.      data/word_edits.jsonl records the before and after of
                                   each field, so a run can resume where it stopped and
                                   any change can be traced or put back by hand.

Three things are held back, and a run says so rather than passing over them quietly:

    The split sheet.        Its rows are senses that need new word entries; editing the
                            shared entry in place is the mistake that sheet exists to
                            prevent.
    Wrong Sense rows.       Re-sensing an entry that several books share strands the
                            books using the sense being replaced. WRONG_SENSE_CLEARED
                            lists the ids checked and found to be on a single book,
                            where there is no other book to strand.
    Entries that would      The database holds a unique constraint on word, part of
    duplicate another.      speech and definition. The save mutation answers HTTP 200
                            with the failure in the body, so a rejected save looks like
                            a successful one; the cached global list is checked first so
                            the request is never spent.

A run outlives its login, and an expired session is indistinguishable from a dead server
-- the page simply never loads. Rather than stopping the whole run, a failed load logs in
again and retries once before counting as a failure.

The Korean and Vietnamese definitions are edited too, but only where the sheet carries a
replacement -- a reworded English definition usually leaves its translations still true,
so most rows change the English alone. The image and the audio are never touched.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright

import rs_scrape as rs
from word_ids import norm

REPORT = Path("references/0820_error-reports_by-type.xlsx")
SHEET = "Vocab Changes"
SPLIT_SHEET = "Vocab Changes - Split"
JOURNAL = Path("data/word_edits.jsonl")
HEADER_ROW = 4

# The form's inputs carry no id or name, so each is found by its placeholder.
PH_ID = "Automatically generated when word is saved"
PH_WORD = "Enter the word you want to add to the system"
PH_DEF = "Enter the definition of the word in English"
PH_KOR = "Enter the definition of the word in Korean"
PH_VIE = "Enter the definition of the word in Vietnamese"
# A div on the page shares this id with the select, so the tag has to be part of it.
POS_SELECT = "select#partOfSpeech"
SAVE_TEXT = "Save Word"

# Saving is heavier on the server than reading, so it gets a longer gap than the scrape.
WRITE_DELAY_S = 5.0
# The inputs appear with their values before the page has wired up its change handlers,
# and text typed into that gap lands in the box without the form ever noticing. Waiting
# is what makes an edit register; the read-back below is what proves it did.
SETTLE_MS = 1500
FIELDS = (("word", "fixed_word", "current_word"),
          ("pos", "fixed_pos", "current_pos"),
          ("definition", "fixed_def", "current_def"),
          ("kor", "fixed_kor", "current_kor"),
          ("vie", "fixed_vie", "current_vie"))
# The English word, part of speech and definition are checked on every page: they came
# from the same scrape as the fixes, so a mismatch means the entry moved on. The
# translations were read separately and are only checked when one is being replaced --
# there is no need to refuse an English fix because somebody retranslated the Korean.
IDENTITY = ("word", "pos", "definition")
# A Wrong Sense fix rewrites what the entry means. The entry is shared, so a book that
# uses the word in the sense being replaced would be left with a definition that no
# longer describes its own story sentence. Those entries need splitting into separate
# word records, not editing in place, so they are held back from every run. The whole
# row is skipped, translations included: those were rewritten to match the new English
# sense, and applying them without it would leave the entry worse than it started.
SKIP_TYPES = ("Wrong Sense",)
# Entries checked and found to be linked to a single book. The reason Wrong Sense is
# held back is that re-sensing a shared entry strands the other books using it -- with
# only one book on the entry there is no other book to strand, so the fix is safe to
# apply in place. Each id below was confirmed against the book list before being added;
# do not extend this by hand without doing the same, because an entry a second book
# picks up later stops being safe.
WRONG_SENSE_CLEARED = {
    "3630", "4054", "33572", "34263", "34532", "34754", "35217", "40203", "40418",
    "41223", "41776", "42029", "42654", "42837", "42882", "42914", "43040", "43110",
    "43411", "43574", "43576", "43577", "43680", "43998", "44317", "44475", "44673",
    "44948", "45312", "45367", "45576", "46654", "47098", "47457", "48233", "49147",
    "49252", "49350", "49505", "49796", "49798", "49799",
}
# Entries the report answered with text that was itself faulty, caught by re-scraping
# the word list and reading back what this script had written. Each is (what went in,
# what it should have said) per field; the run expects the first on the page, so an
# amendment applies once and is a no-op afterwards. They bypass the journal, whose whole
# record of these words is that they are finished.
#
# 40121 is the one that matters: the report's fix cell began "Change the definition to:"
# and the whole string was written, so the instruction became the definition every book
# linked to that entry showed.
AMENDED = {
    "40121": {"definition": (
        "Change the definition to: a woman who is married to a child’s father but "
        "is not the child’s birth mother.",
        "a woman who is married to a child’s father but is not the child’s "
        "birth mother")},
}

PLACEHOLDER = {"word": PH_WORD, "definition": PH_DEF, "kor": PH_KOR, "vie": PH_VIE}


# --------------------------------------------------------------------------------------
# The sheet
# --------------------------------------------------------------------------------------


def load_rows(path: Path, sheet: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"{path} has no sheet named {sheet!r}")
    ws = wb[sheet]
    hdr = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)
           if ws.cell(HEADER_ROW, c).value}
    # "Error Types" belongs here even though nothing is read out of it directly: the
    # Wrong Sense hold is driven by it, and a lookup for a column that is absent returns
    # empty, so without this a report missing the column would quietly write every entry
    # the hold exists to keep back.
    needed = {"word_id", "current_word", "fixed_word", "current_pos", "fixed_pos",
              "current_def", "fixed_def", "Error Types"}
    missing = needed - set(hdr)
    if missing:
        sys.exit(f"{sheet} is missing column(s): {', '.join(sorted(missing))}")
    # A report generated before the translations were scraped simply has no such
    # columns, and then no translation is touched.
    if "current_kor" not in hdr:
        print(f"  ! {sheet} has no translation columns -- Korean and Vietnamese will "
              f"be left alone")

    # Anything on the split sheet must never be edited in place.
    split_ids = set()
    if SPLIT_SHEET in wb.sheetnames:
        sp = wb[SPLIT_SHEET]
        col = {sp.cell(HEADER_ROW, c).value: c for c in range(1, sp.max_column + 1)
               if sp.cell(HEADER_ROW, c).value}.get("word_id")
        for r in range(HEADER_ROW + 1, sp.max_row + 1):
            v = sp.cell(r, col).value if col else None
            if v:
                split_ids.add(str(v).rstrip("ABCDEFGHIJ"))

    rows, held = [], 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        g = lambda name: (str(ws.cell(r, hdr[name]).value or "").strip()
                          if name in hdr else "")
        wid = g("word_id")
        if not wid:
            continue
        if not wid.isdigit():
            print(f"  ! row {r}: word_id {wid!r} is not a plain id -- skipped")
            continue
        if wid in split_ids:
            print(f"  ! row {r}: word {wid} is a split entry -- skipped")
            continue
        types = [t.strip() for t in g("Error Types").split(",") if t.strip()]
        if any(t in SKIP_TYPES for t in types) and wid not in WRONG_SENSE_CLEARED:
            held += 1
            continue
        row = {"row": r, "word_id": wid, "types": types}
        for name, fixed, current in FIELDS:
            row[f"fixed_{name}"] = g(fixed)
            row[f"current_{name}"] = g(current)
        amend = AMENDED.get(wid)
        if amend:
            # Both halves are replaced: the page holds what this script last wrote
            # there, not what the report quoted before any of it ran.
            for name, (was, should) in amend.items():
                row[f"current_{name}"] = was
                row[f"fixed_{name}"] = should
            for name, _, _ in FIELDS:
                if name not in amend:
                    row[f"fixed_{name}"] = ""
            row["amended"] = True
        if any(row[f"fixed_{n}"] for n, _, _ in FIELDS):
            rows.append(row)
    if held:
        print(f"  held back {held} row(s) whose reason includes "
              f"{' or '.join(SKIP_TYPES)}")
    return rows


def saved_fields() -> dict[str, dict[str, str]]:
    """word id -> the field values the journal records as already written."""
    out: dict[str, dict[str, str]] = {}
    if not JOURNAL.exists():
        return out
    for line in JOURNAL.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        if rec.get("status") != "saved":
            continue
        entry = out.setdefault(str(rec["word_id"]), {})
        for f in rec.get("fields", []):
            entry[f] = rec["after"][f]
    return out


def already_done(row: dict, done: dict[str, dict[str, str]]) -> bool:
    """True only when every field this row wants to change is already written.

    Asking merely whether the word has been saved before is not enough: a word edited
    in an earlier run for its definition would then be treated as finished, and a
    translation still pending on the same word would be skipped for good. That is
    exactly what happened to three entries whose Vietnamese never went in.
    """
    have = done.get(row["word_id"])
    if not have:
        return False
    for name, _, _ in FIELDS:
        want = row[f"fixed_{name}"]
        if want and norm(have.get(name, "")) != norm(want):
            return False
    return True


def journal(rec: dict) -> None:
    JOURNAL.parent.mkdir(parents=True, exist_ok=True)
    with JOURNAL.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(rec, ensure_ascii=False) + "\n")


# --------------------------------------------------------------------------------------
# The page
# --------------------------------------------------------------------------------------


def collision_index() -> dict[tuple, list[str]]:
    """(word, part of speech, English definition) -> the ids already using it.

    The database holds a unique constraint on those three columns, so an edit that
    lands on a combination another entry already has is rejected -- the mutation
    returns HTTP 200 with the error in the body, which looks like success. Checking
    against the cached global list first turns that into a clean skip, and the cache
    costs nothing to read.
    """
    import word_ids

    if not word_ids.CACHE.exists():
        print("  ! no global word cache -- cannot check for duplicate entries first")
        return {}
    out: dict[tuple, list[str]] = {}
    for g in word_ids.load_global():
        out.setdefault(_key(g["word"], g["pos"], g["definition"]), []).append(g["id"])
    return out


def _key(word: str, pos: str, definition: str) -> tuple:
    """The three columns the constraint covers, compared the way the database compares
    them -- case and all. Lowercasing here blocks edits the database would have
    accepted: `bonnet(s)` and `Bonnet(s)` are two rows to Postgres, not one."""
    import word_ids

    return (word_ids.norm(word), word_ids.pos_key(pos), word_ids.norm(definition))


def collides(row: dict, index: dict) -> list[str]:
    import word_ids

    key = _key(row["fixed_word"] or row["current_word"],
               row["fixed_pos"] or row["current_pos"],
               row["fixed_definition"] or row["current_definition"])
    return [i for i in index.get(key, []) if i != row["word_id"]]


def field(page, placeholder):
    return page.get_by_placeholder(placeholder, exact=False).first


def read_form(page) -> dict:
    out = {"word_id": rs._clean(field(page, PH_ID).input_value()),
           "pos": rs._clean(page.locator(POS_SELECT).input_value())}
    for name, placeholder in PLACEHOLDER.items():
        out[name] = rs._clean(field(page, placeholder).input_value())
    return out


def save_button(page):
    return page.locator("button", has_text=SAVE_TEXT).first


def check_matches(row: dict, actual: dict) -> str | None:
    """None when the form holds what the report described, else why it does not."""
    if actual["word_id"] != row["word_id"]:
        return (f"the form is word {actual['word_id']!r}, not {row['word_id']!r}")
    for name, _, _ in FIELDS:
        if name not in IDENTITY and not row[f"fixed_{name}"]:
            continue
        want, got = row[f"current_{name}"], actual[name]
        if norm(want) != norm(got):
            return (f"{name} on the site is {got!r}, but the report recorded "
                    f"{want!r} -- it has been edited since")
    return None


def pos_options(page) -> set[str]:
    return {o.get_attribute("value") or ""
            for o in page.locator(f"{POS_SELECT} option").all()}


def open_word(page, url: str) -> None:
    rs.polite_goto(page, url)
    page.wait_for_selector(f"[placeholder='{PH_WORD}']",
                           timeout=rs.SELECTOR_TIMEOUT_MS)
    page.wait_for_timeout(SETTLE_MS)


def apply_edits(page, row: dict) -> tuple[list[str], list[str]]:
    """Fill the changed fields. Returns (fields changed, problems)."""
    changed, problems = [], []
    if row["fixed_word"]:
        box = field(page, PH_WORD)
        box.fill("")
        box.fill(row["fixed_word"])
        changed.append("word")
    if row["fixed_pos"]:
        want = row["fixed_pos"].strip().lower()
        if want not in pos_options(page):
            problems.append(f"no part-of-speech option named {row['fixed_pos']!r}")
        else:
            page.locator(POS_SELECT).select_option(want)
            changed.append("pos")
    for name in ("definition", "kor", "vie"):
        if row[f"fixed_{name}"]:
            box = field(page, PLACEHOLDER[name])
            box.fill("")
            box.fill(row[f"fixed_{name}"])
            changed.append(name)
    return changed, problems


def wanted(row: dict, actual: dict) -> dict:
    """What the record should read once the row has been applied."""
    return {n: row[f"fixed_{n}"] or actual[n] for n, _, _ in FIELDS}


# --------------------------------------------------------------------------------------
# The run
# --------------------------------------------------------------------------------------


def run(args) -> None:
    sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
    rows = load_rows(args.report, args.sheet)
    if args.only:
        keep = {s.strip() for s in args.only.split(",") if s.strip()}
        rows = [r for r in rows if r["word_id"] in keep]
    if not args.redo:
        done = saved_fields()
        before = len(rows)
        # An amendment exists because what went in was wrong, and the journal records
        # only that it went in, so it must survive the resume filter.
        rows = [r for r in rows if r.get("amended") or not already_done(r, done)]
        if before - len(rows):
            print(f"{before - len(rows)} word(s) already have every pending change "
                  f"written -- skipping")
    if args.limit:
        rows = rows[:args.limit]
    if not rows:
        sys.exit("nothing to do")

    mode = "APPLYING" if args.apply else "DRY RUN (nothing will be saved)"
    print(f"{mode}: {len(rows)} word(s), about "
          f"{len(rows) * (WRITE_DELAY_S + 3) / 60:.0f} min at current pacing\n")

    counts = {"saved": 0, "would save": 0, "skipped": 0, "failed": 0}
    consecutive = 0
    index = collision_index()
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=rs.HEADLESS)
        state = str(rs.STORAGE_STATE) if rs.STORAGE_STATE.exists() else None
        context = browser.new_context(storage_state=state)
        rs.block_heavy_resources(context)
        page = context.new_page()
        page.set_default_navigation_timeout(rs.NAV_TIMEOUT_MS)
        rs.ensure_logged_in(page, *rs.credentials())
        rs.save_session(context)

        try:
            for n, row in enumerate(rows, 1):
                wid = row["word_id"]
                label = f"[{n}/{len(rows)}] {wid}"
                url = f"{rs.BASE_URL}/words/{wid}/edit"
                # Checked before the page is fetched: a row the database will refuse
                # costs nothing to rule out, and loading its page would be a wasted
                # request.
                clash = collides(row, index)
                if not clash:
                    try:
                        open_word(page, url)
                    except (PlaywrightTimeout, PlaywrightError):
                        # A run long enough outlives its login, and an expired session
                        # looks exactly like a dead server: the page never finishes
                        # loading. Log in again and try once more before believing the
                        # server is in trouble, or a whole run stops for a session that
                        # merely aged out.
                        print(f"{label}: page did not load -- logging in again")
                        try:
                            rs.ensure_logged_in(page, *rs.credentials())
                            rs.save_session(context)
                            open_word(page, url)
                        except (PlaywrightTimeout, PlaywrightError) as exc:
                            consecutive += 1
                            counts["failed"] += 1
                            print(f"{label}: {type(exc).__name__} loading the page")
                            journal({"word_id": wid, "status": "failed",
                                     "why": f"{type(exc).__name__} loading the page"})
                            if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                                print(f"\n{consecutive} failures in a row -- stopping.")
                                break
                            continue

                if clash:
                    consecutive = 0
                    counts["skipped"] += 1
                    why = (f"another entry (#{clash[0]}) already has this word, part of "
                           f"speech and definition, and the database will not allow two")
                    print(f"{label}: SKIPPED -- {why}")
                    journal({"word_id": wid, "status": "skipped", "why": why,
                             "duplicate_of": clash})
                    continue

                actual = read_form(page)
                why = check_matches(row, actual)
                if why:
                    consecutive = 0
                    counts["skipped"] += 1
                    print(f"{label}: SKIPPED -- {why}")
                    journal({"word_id": wid, "status": "skipped", "why": why,
                             "on_site": actual})
                    continue

                changed, problems = apply_edits(page, row)
                # Read the form back. A fill that arrived before the page was ready sits
                # in the box unnoticed, and two fields do not always take together, so
                # checking each one is what rules out saving a half-applied edit.
                target = wanted(row, actual)
                page.wait_for_timeout(400)
                stuck = [f for f in changed
                         if norm(read_form(page)[f]) != norm(target[f])]
                if stuck:
                    apply_edits(page, row)
                    page.wait_for_timeout(800)
                    stuck = [f for f in changed
                             if norm(read_form(page)[f]) != norm(target[f])]
                if stuck:
                    consecutive = 0
                    counts["skipped"] += 1
                    why = f"{', '.join(stuck)} would not take in the form"
                    print(f"{label}: SKIPPED -- {why}")
                    journal({"word_id": wid, "status": "skipped", "why": why,
                             "on_site": actual})
                    continue
                for p in problems:
                    print(f"{label}: ! {p}")
                if not changed:
                    consecutive = 0
                    counts["skipped"] += 1
                    journal({"word_id": wid, "status": "skipped",
                             "why": "no usable field to change", "on_site": actual})
                    print(f"{label}: SKIPPED -- nothing usable to change")
                    continue

                # The form enables Save only once a value really differs; if it is still
                # disabled the edits did not register and saving would be a no-op.
                if save_button(page).is_disabled():
                    consecutive = 0
                    counts["skipped"] += 1
                    print(f"{label}: SKIPPED -- the form says nothing changed")
                    journal({"word_id": wid, "status": "skipped",
                             "why": "save stayed disabled after filling",
                             "on_site": actual, "fields": changed})
                    continue

                rec = {"word_id": wid, "fields": changed, "before": actual,
                       "after": target, "problems": problems}
                if not args.apply:
                    consecutive = 0
                    counts["would save"] += 1
                    print(f"{label}: would change {', '.join(changed)}")
                    for f in changed:
                        print(f"        {f}: {actual[f]!r}")
                        print(f"        {' ' * len(f)}  -> {rec['after'][f]!r}")
                    continue

                try:
                    save_button(page).click()
                    page.wait_for_timeout(1500)
                except (PlaywrightTimeout, PlaywrightError) as exc:
                    consecutive += 1
                    counts["failed"] += 1
                    print(f"{label}: {type(exc).__name__} clicking save")
                    journal({**rec, "status": "failed",
                             "why": f"{type(exc).__name__} clicking save"})
                    if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                        print(f"\n{consecutive} failures in a row -- stopping.")
                        break
                    continue

                if args.verify:
                    # The save has already gone through at this point, so a failure to
                    # reload is only a failure to confirm. It is recorded as unverified
                    # and the run carries on; letting it raise would abandon the whole
                    # run over a page that did not come back.
                    try:
                        open_word(page, url)
                    except (PlaywrightTimeout, PlaywrightError):
                        try:
                            rs.ensure_logged_in(page, *rs.credentials())
                            rs.save_session(context)
                            open_word(page, url)
                        except (PlaywrightTimeout, PlaywrightError) as exc:
                            consecutive += 1
                            counts["saved"] += 1
                            print(f"{label}: saved {', '.join(changed)}, but could not "
                                  f"reload to confirm ({type(exc).__name__})")
                            journal({**rec, "status": "saved",
                                     "why": "saved but not verified: "
                                            f"{type(exc).__name__} on reload"})
                            if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                                print(f"\n{consecutive} failures in a row -- stopping.")
                                break
                            time.sleep(WRITE_DELAY_S)
                            continue
                    after = read_form(page)
                    wrong = [f for f in changed
                             if norm(after[f]) != norm(rec["after"][f])]
                    if wrong:
                        consecutive += 1
                        counts["failed"] += 1
                        print(f"{label}: SAVE DID NOT STICK for {', '.join(wrong)}")
                        journal({**rec, "status": "failed",
                                 "why": "reload did not show the new value",
                                 "on_reload": after})
                        if consecutive >= rs.MAX_CONSECUTIVE_FAILURES:
                            print(f"\n{consecutive} failures in a row -- stopping.")
                            break
                        continue
                    rec["on_reload"] = after

                consecutive = 0
                counts["saved"] += 1
                journal({**rec, "status": "saved"})
                print(f"{label}: saved {', '.join(changed)}")
                time.sleep(WRITE_DELAY_S)
        finally:
            browser.close()

    print("\n" + "  ".join(f"{k}: {v}" for k, v in counts.items()))
    if not args.apply:
        print(f"\nDRY RUN -- nothing was saved. Add --apply to write these changes.")
    else:
        print(f"journal: {JOURNAL}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--apply", action="store_true",
                    help="actually save; without it the run only reports")
    ap.add_argument("--report", type=Path, default=REPORT)
    ap.add_argument("--sheet", default=SHEET)
    ap.add_argument("--limit", type=int, help="only handle this many words")
    ap.add_argument("--only", help="comma-separated word ids to handle")
    ap.add_argument("--redo", action="store_true",
                    help="handle words whose pending changes the journal already "
                         "records as written")
    ap.add_argument("--no-verify", dest="verify", action="store_false",
                    help="skip the reload that confirms each save stuck")
    args = ap.parse_args()
    if args.sheet == SPLIT_SHEET:
        sys.exit(f"{SPLIT_SHEET!r} holds senses that need new word entries, not edits "
                 f"to existing ones. This script will not touch it.")
    run(args)


if __name__ == "__main__":
    main()
